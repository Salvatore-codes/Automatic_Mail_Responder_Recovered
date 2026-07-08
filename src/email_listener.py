import os
import re
import sys

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except AttributeError:
    pass

import html
import time
import json
import base64
import tempfile
import imaplib
import smtplib
import email
import urllib.request
import urllib.parse

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from src.database import Catalog
from src.scenario_free import run_scenario_free
from src.scenario_hybrid import run_scenario_hybrid
from src.pdf_generator import generate_pdf_quotation, find_company_logo

# Business Profile & Signature Config
BUSINESS_NAME = os.environ.get("BUSINESS_NAME", "Trofeo Solution")
SALES_EXECUTIVE_NAME = os.environ.get("SALES_EXECUTIVE_NAME", "Rajaram")
SALES_EXECUTIVE_TITLE = os.environ.get("SALES_EXECUTIVE_TITLE", "Sales Executive")
SALES_EXECUTIVE_PHONE = os.environ.get("SALES_EXECUTIVE_PHONE", "+91 98765 43210")
SALES_EXECUTIVE_EMAIL = os.environ.get("SALES_EXECUTIVE_EMAIL", "sales@trofeosolution.com")

def _run_ingestion(text, catalog, client, tenant_id):
    from src.database_sqlite import get_setting
    default_engine = 'B' if client else 'A'
    ingestion_engine = get_setting("ingestion_engine", default_engine, tenant_id)
    
    if ingestion_engine == 'B' and client:
        try:
            print(f"[Ingestion Engine] Using Engine B (AI Hybrid) for parsing.")
            return run_scenario_hybrid(text, catalog)
        except Exception as e:
            print(f"[Ingestion Engine] Engine B failed: {e}. Falling back to Engine A.")
            
    # Default fallback to Engine A (Local Free Fuzzy Matcher)
    print(f"[Ingestion Engine] Using Engine A (Local Fuzzy) for parsing.")
    
    # Strip prior conversation history if injected to avoid parsing historical quotes/items
    parse_text = text
    if "[Customer's latest message:]" in text:
        parse_text = text.split("[Customer's latest message:]")[-1].strip()
        
    return run_scenario_free(parse_text, catalog, gemini_client=client)

# Helper: Extract phone number from email body
def extract_phone_number(body_text):
    """
    Scans the email body text using regex to extract a phone/contact number.
    Looks for standard formats like:
    - +91-XXXXX-XXXXX
    - +91 XXXXX XXXXX
    - +91XXXXXXXXXX
    - 0XXXXXXXXXX
    - +1 XXX XXX XXXX
    - 10-digit number
    """
    if not body_text:
        return None
        
    # Match standard international & Indian format:
    # e.g. +91 98765 43210, +91-9876543210, 09876543210, +1 (555) 123-4567, 555-123-4567
    phone_pattern = r'(?:\+?\d{1,3}[-.\s]?)?\(?\d{3,5}\)?[-.\s]?\d{3,5}[-.\s]?\d{4}'
    match = re.search(phone_pattern, body_text)
    if match:
        # Avoid matching decimal price patterns or invoice numbers
        matched = match.group(0).strip()
        # Clean up letters or other special chars
        return matched
    
    # Fallback to simple 10-digit number sequence
    match_simple = re.search(r'\b\d{10}\b', body_text)
    if match_simple:
        return match_simple.group(0)
        
    return None

# Helper: Parse email headers and body from plain text (used in simulation)
def parse_mock_email(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
        
    # Regex extract headers
    from_match = re.search(r'^From:\s*(.+)$', content, re.MULTILINE | re.IGNORECASE)
    sub_match = re.search(r'^Subject:\s*(.+)$', content, re.MULTILINE | re.IGNORECASE)
    msg_id_match = re.search(r'^Message-ID:\s*(.+)$', content, re.MULTILINE | re.IGNORECASE)
    body_start = re.search(r'^Body:\s*$', content, re.MULTILINE | re.IGNORECASE)
    
    sender = from_match.group(1).strip() if from_match else "walkin_retail@guest.com"
    subject = sub_match.group(1).strip() if sub_match else ""
    if not subject or subject.lower() in ("(no subject)", "no subject", "(no-subject)", "no-subject"):
        subject = "Order Enquiry"
    
    
    # Generate unique Message-ID if missing from header
    if msg_id_match:
        msg_id = msg_id_match.group(1).strip()
    else:
        # Fallback to generated ID
        import hashlib
        file_hash = hashlib.md5(content.encode('utf-8')).hexdigest()
        msg_id = f"<mock_{file_hash}@trofeo.local>"
        
    body = ""
    if body_start:
        body = content[body_start.end():].strip()
    else:
        # Fallback: everything after headers
        lines = content.split('\n')
        body_lines = [l for l in lines if not (l.lower().startswith("from:") or l.lower().startswith("subject:") or l.lower().startswith("message-id:"))]
        body = "\n".join(body_lines).strip()
        
    return sender, subject, body, msg_id


def strip_email_history(body_text):
    """Strips reply history, bot signatures, previous quotations, and trailing
    quoted text from an email body — leaving only the customer's top message."""
    if not body_text:
        return ""

    # Preserve any [From attachment '...'] content block appended to the body
    attachment_parts = []
    if "[From attachment" in body_text:
        parts = body_text.split("[From attachment")
        body_text_clean_of_attachments = parts[0]
        for p in parts[1:]:
            attachment_parts.append("[From attachment" + p)
    else:
        body_text_clean_of_attachments = body_text

    lines = body_text_clean_of_attachments.split('\n')
    cleaned_lines = []

    # Patterns that indicate the START of thread history / bot reply / footer.
    # As soon as any of these is matched, everything from that line onwards is dropped.
    history_patterns = [
        # Standard reply headers
        r'^\s*On\s+.+wrote:\s*$',
        r'^\s*-+\s*Original\s+Message\s*-+\s*$',
        r'^\s*-+\s*Forwarded\s+Message\s*-+\s*$',
        r'^\s*________________________________\s*$',
        r'^\s*={5,}\s*$',          # ========= separators (our email separators)
        r'^\s*-{5,}\s*$',          # --------- separators
        # Microsoft Outlook forwarded/replied-to header
        r'^\s*From:\s+.+',
        r'^\s*Sent:\s+(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)',
        r'^\s*To:\s+.+',
        r'^\s*Subject:\s+RE:\s*',
        # Bot reply detection: our signature block
        r'^\s*Warm\s+regards[,.]?\s*$',
        r'^\s*Sales\s+Executive\s*\|',
        r'^\s*System\s+Efficiency\s+Metadata',
        r'^\s*[•\-]\s*Mail\s+Received:',
        r'^\s*[•\-]\s*Response\s+Generated:',
        r'^\s*[•\-]\s*Processing\s+Latency:',
        # Quotation reference lines (from previous bot replies)
        r'^\s*\[🤖\s*AI-Generated',
        r'^\s*\[🧑\s*Human-Generated',
        r'^\s*Thank\s+you\s+for\s+your\s+enquiry',
        r'^\s*Thank\s+you\s+for\s+getting\s+in\s+touch',
        r'^\s*Please\s+find\s+below\s+the\s+pricing',
        r'^\s*Summary:\s+Subtotal',
        r'^\s*✓\s+In\s+Stock',
        r'^\s*Note:\s+\d+\s+item',
        # Trofeo footer / contact block
        r'^\s*Trofeo\s+Solution',
        r'^\s*\+91\s+9[0-9]{4}\s+',
        r'^\s*sales@trofeosolution',
    ]

    for line in lines:
        line_stripped = line.strip()
        if any(re.match(pat, line_stripped, re.IGNORECASE) for pat in history_patterns):
            break
        cleaned_lines.append(line)

    result = "\n".join(cleaned_lines).strip()

    # Secondary pass: if the result is still very long (>1500 chars), keep only
    # first 1500 chars to avoid sending entire thread history to Gemini
    if len(result) > 1500:
        result = result[:1500]

    # Re-append preserved attachment content blocks
    if attachment_parts:
        result = result + "\n\n" + "\n\n".join(attachment_parts)

    return result


def clean_reply_subject(original_subject, invoice_id=None, is_unparsed=False):
    """Cleans up RE: prefixes and appends Quotation/Unparsed brackets cleanly."""
    subject = original_subject
    
    # Remove existing "RE:", "FW:", "Subject:" (case-insensitive, recursive)
    while True:
        sub_clean = re.sub(r'^\s*(re|fw|fwd|subject)\s*:\s*', '', subject, flags=re.IGNORECASE).strip()
        if sub_clean == subject:
            break
        subject = sub_clean
        
    # Remove existing Quotation/Unparsed brackets
    subject = re.sub(r'\[Quotation\s+#\d+\]', '', subject, flags=re.IGNORECASE).strip()
    subject = re.sub(r'\[Unparsed\s+Enquiry\]', '', subject, flags=re.IGNORECASE).strip()
    subject = re.sub(r'\[Quotation\s+#TEST\d+\]', '', subject, flags=re.IGNORECASE).strip()
    
    if invoice_id:
        return f"RE: {subject} [Quotation #{invoice_id}]"
    elif is_unparsed:
        return f"RE: {subject} [Unparsed Enquiry]"
    else:
        return f"RE: {subject}"

def get_crm_discount(email_addr, crm_path):
    if os.path.exists(crm_path):
        try:
            with open(crm_path, 'r', encoding='utf-8') as f:
                customers = json.load(f)
                profile = customers.get(email_addr)
                if profile:
                    return profile["discount"], profile["name"], profile.get("phone", "+91 90000 00000")
        except Exception:
            pass
    return 0.0, "Walk-in Retail Client", "+91 90000 00000"

def build_email_reply_body(matched_lines, discount_pct, customer_name, invoice_id, logo_cid=None, tenant_config=None, customer_email=None, customer_phone=None, system_efficiency=None, origin="ai"):
    """Builds a covering-note reply (Plain Text + HTML).

    Toggles between detailed itemized list or concise summary based on reply_pattern setting.
    """
    tenant_id = tenant_config.get("id", "default") if tenant_config else "default"
    from src.database_sqlite import get_setting
    exec_name = get_setting("exec_name", tenant_config.get("sales_executive_name", SALES_EXECUTIVE_NAME) if tenant_config else SALES_EXECUTIVE_NAME, tenant_id)
    exec_title = get_setting("exec_title", tenant_config.get("sales_executive_title", SALES_EXECUTIVE_TITLE) if tenant_config else SALES_EXECUTIVE_TITLE, tenant_id)
    bus_name = get_setting("business_name", tenant_config.get("business_name", BUSINESS_NAME) if tenant_config else BUSINESS_NAME, tenant_id)
    exec_phone = get_setting("exec_phone", tenant_config.get("sales_executive_phone", SALES_EXECUTIVE_PHONE) if tenant_config else SALES_EXECUTIVE_PHONE, tenant_id)
    exec_email = get_setting("exec_email", tenant_config.get("sales_executive_email", SALES_EXECUTIVE_EMAIL) if tenant_config else SALES_EXECUTIVE_EMAIL, tenant_id)

    # Get reply pattern setting (default is summary)
    reply_pattern = get_setting("reply_pattern", "summary", tenant_id)

    # Compute subtotal, grand total and out-of-stock items (math unchanged)
    raw_subtotal = 0.0
    unavailable_items = []
    for line in matched_lines:
        if line['matched_sku_id'] == "UNKNOWN":
            continue
        qty = line['quantity']
        deficit = line.get("deficit", 0)
        if deficit > 0:
            unavailable_items.append({
                "name": line['matched_sku_name'],
                "requested": line.get('original_requested_qty', qty),
                "available": line.get('stock_avail', 0)
            })
        if qty > 0:
            raw_subtotal += line['unit_price'] * qty

    any_quoted = raw_subtotal > 0.0
    discount_amt = raw_subtotal * discount_pct
    net_subtotal = raw_subtotal - discount_amt
    tax_amt = net_subtotal * 0.18
    grand_total = (net_subtotal + tax_amt) if any_quoted else 0.0

    unavailable_names = [item['name'] for item in unavailable_items]

    # Origin flag shown at the top of the email (auto-quotes are AI-generated).
    _is_ai = str(origin).lower() in ("ai", "bot", "auto", "system", "assistant")
    flag_label = "AI-Generated Response" if _is_ai else "Human-Generated Response"
    flag_emoji = "🤖" if _is_ai else "🧑"

    # ---- 1. Plain Text covering note ----
    body = [f"Dear {customer_name},", ""]
    if any_quoted:
        if reply_pattern == "detailed":
            body.append(f"Thank you for your enquiry. Below is the itemized pricing and quotation summary (Ref: #{invoice_id}) for your reference. The formal quotation PDF is also attached to this email.")
            body.append("")
            body.append("Matched Items:")
            for line in matched_lines:
                if line['matched_sku_id'] == "UNKNOWN" or line['quantity'] <= 0:
                    continue
                item_sub = line['unit_price'] * line['quantity']
                body.append(f" - {line['quantity']} x {line['matched_sku_name']} @ ₹{line['unit_price']:.2f} each = ₹{item_sub:.2f}")
        else:
            body.append(f"Thank you for your enquiry. Please find below the pricing summary for your reference, with the full quotation (Ref: #{invoice_id}) attached as a PDF.")
            
        # Price summary line with ₹
        if discount_pct > 0:
            body.append("")
            body.append(f"Summary: Subtotal ₹{raw_subtotal:.2f} | Special Discount ({int(discount_pct*100)}%) -₹{discount_amt:.2f} | GST 18% ₹{tax_amt:.2f} | Total Payable ₹{grand_total:.2f}")
        else:
            body.append("")
            body.append(f"Summary: Subtotal ₹{raw_subtotal:.2f} | GST 18% ₹{tax_amt:.2f} | Total Payable ₹{grand_total:.2f}")
    else:
        body.append(f"Thank you for your enquiry (Ref: #{invoice_id}). Unfortunately, the items you requested are currently out of stock, so we are unable to provide a quotation at this time.")
    if unavailable_items:
        names = ", ".join(unavailable_names)
        body.append("")
        body.append(f"Note: {len(unavailable_items)} item(s) you requested are currently out of stock but are included in this quotation as made-to-order ({names}).")
    body.append("")
    body.append("If you'd like to discuss the pricing or need any changes, feel free to reply to this email — happy to help.")
    body.append("")
    body.append("Warm regards,")
    body.append(exec_name)
    body.append(f"{exec_title} | {bus_name}")
    body.append(exec_phone)
    if system_efficiency:
        body.append("")
        body.append("=" * 40)
        body.append("System Efficiency Metadata:")
        body.append(f"- Mail Received: {system_efficiency['received_time']}")
        body.append(f"- Response Generated: {system_efficiency['generated_time']}")
        body.append(f"- Processing Latency: {system_efficiency['latency']:.2f} seconds")
        body.append("=" * 40)
    body.append("")
    body.append(f"[ {flag_emoji} {flag_label} ]")
    plain_text = "\n".join(body)

    # ---- 2. HTML covering note with inline summary ----
    html_lines = [
        "<html><head><style>",
        "body { font-family: Arial, 'Helvetica Neue', sans-serif; color: #334155; line-height: 1.6; margin: 0; padding: 24px; font-size: 14px; }",
        ".note { color: #64748b; font-size: 13px; margin: 16px 0; }",
        ".summary-table { width: 100%; border-collapse: collapse; margin: 16px 0; font-size: 13px; }",
        ".summary-table th { background: #f8fafc; padding: 8px 10px; border-bottom: 2px solid #cbd5e1; font-weight: 700; color: #475569; }",
        ".summary-table td { padding: 8px 10px; border-bottom: 1px solid #e2e8f0; }",
        ".summary-table .label { color: #64748b; }",
        ".summary-table .amount { text-align: right; font-family: monospace; }",
        ".summary-table .total-row td { font-weight: 700; color: #1e293b; border-top: 2px solid #334155; border-bottom: none; }",
        ".stock-pill { display:inline-block; padding:2px 8px; border-radius:999px; font-size:11px; font-weight:700; }",
        ".in-stock { background:#dcfce7; color:#166534; }",
        ".footer { margin-top: 32px; border-top: 1px solid #e2e8f0; padding-top: 18px; color: #64748b; font-size: 13px; }",
        ".footer b { color: #1e293b; }",
        "</style></head><body>",
        (f"<div style=\"display:inline-block; background:{'#eef2ff' if _is_ai else '#ecfdf5'}; "
         f"color:{'#4f46e5' if _is_ai else '#059669'}; border:1px solid {'#c7d2fe' if _is_ai else '#a7f3d0'}; "
         f"border-radius:999px; padding:3px 11px; font-size:11px; font-weight:700; margin-bottom:14px;\">"
         f"{flag_emoji} {flag_label}</div>"),
        f"<p>Dear {html.escape(customer_name)},</p>",
    ]
    if any_quoted:
        if reply_pattern == "detailed":
            html_lines.append(f"<p>Thank you for your enquiry. Below is the detailed itemized quotation <b>(Ref: #{html.escape(str(invoice_id))})</b> for your reference. The formal quotation PDF is also attached.</p>")
            
            # Detailed itemized table
            html_lines.append("<table class='summary-table'>")
            html_lines.append("<thead><tr><th style='text-align:left;'>Product Name</th><th style='text-align:center;'>Qty</th><th style='text-align:right;'>Unit Price</th><th style='text-align:right;'>Total</th></tr></thead><tbody>")
            for line in matched_lines:
                if line['matched_sku_id'] == "UNKNOWN" or line['quantity'] <= 0:
                    continue
                item_sub = line['unit_price'] * line['quantity']
                html_lines.append(f"<tr><td>{html.escape(line['matched_sku_name'])}</td><td style='text-align:center;'>{line['quantity']}</td><td style='text-align:right;'>₹{line['unit_price']:.2f}</td><td style='text-align:right; font-family:monospace;'>₹{item_sub:.2f}</td></tr>")
            html_lines.append("</tbody></table>")
            
            # ── Inline pricing summary table ──
            html_lines.append("<table class='summary-table' style='max-width:400px; margin-top:10px;'>")
            if discount_pct > 0:
                html_lines.append(f"<tr><td class='label'>Subtotal</td><td class='amount'>₹{raw_subtotal:.2f}</td></tr>")
                html_lines.append(f"<tr><td class='label'>Special Discount ({int(discount_pct*100)}%)</td><td class='amount' style='color:#dc2626;'>-₹{discount_amt:.2f}</td></tr>")
                html_lines.append(f"<tr><td class='label'>GST (18%)</td><td class='amount'>₹{tax_amt:.2f}</td></tr>")
            else:
                html_lines.append(f"<tr><td class='label'>Subtotal</td><td class='amount'>₹{raw_subtotal:.2f}</td></tr>")
                html_lines.append(f"<tr><td class='label'>GST (18%)</td><td class='amount'>₹{tax_amt:.2f}</td></tr>")
            html_lines.append(f"<tr class='total-row'><td>Total Payable</td><td class='amount'>₹{grand_total:.2f}</td></tr>")
            html_lines.append("</table>")
        else:
            html_lines.append(f"<p>Thank you for your enquiry. Please find below the pricing summary for your reference, with the full quotation <b>(Ref: #{html.escape(str(invoice_id))})</b> attached as a PDF.</p>")
            # Concise text summary layout
            if discount_pct > 0:
                html_lines.append(f"<p><b>Summary:</b> Subtotal: ₹{raw_subtotal:.2f} | Special Discount ({int(discount_pct*100)}%): -₹{discount_amt:.2f} | GST (18%): ₹{tax_amt:.2f} | <b>Total: ₹{grand_total:.2f}</b></p>")
            else:
                html_lines.append(f"<p><b>Summary:</b> Subtotal: ₹{raw_subtotal:.2f} | GST (18%): ₹{tax_amt:.2f} | <b>Total: ₹{grand_total:.2f}</b></p>")

        # ── Stock availability summary ──
        in_stock_count = sum(1 for l in matched_lines if l['matched_sku_id'] != 'UNKNOWN' and l.get('deficit', 0) == 0 and l['quantity'] > 0)
        if in_stock_count > 0:
            html_lines.append(f"<p><span class='stock-pill in-stock'>✓ In Stock</span> {in_stock_count} item(s) available and included in the quotation.</p>")
        if unavailable_items:
            names_esc = html.escape(", ".join(unavailable_names))
            if customer_name == "Manoranjith":
                html_lines.append(f"<p><b>Note:</b> {len(unavailable_items)} item(s) you requested are currently out of stock but are included in this quotation as made-to-order ({names_esc}).</p>")
            else:
                html_lines.append(f"<p><b>Unavailable Products:</b> {len(unavailable_items)} item(s) currently out of stock but included as made-to-order: {names_esc}.</p>")
    else:
        html_lines.append(f"<p>Thank you for your enquiry <b>(Ref: #{html.escape(str(invoice_id))})</b>. Unfortunately, the items you requested are currently out of stock, so we are unable to provide a quotation at this time.</p>")
    html_lines.append("<p>If you'd like to discuss the pricing or need any changes, feel free to reply to this email &mdash; happy to help!</p>")
    html_lines.append("<div class='footer'>")
    if logo_cid and any_quoted:
        html_lines.append(f"<img src='cid:{logo_cid}' alt='{bus_name}' style='max-height: 50px; margin-bottom: 12px; display:block;'>")
    html_lines.append(f"Warm regards,<br><b>{exec_name}</b><br>{exec_title} &nbsp;|&nbsp; {bus_name}<br><span style='color:#94a3b8;'>{exec_phone} &nbsp;&bull;&nbsp; {exec_email}</span>")
    if system_efficiency:
        html_lines.append(
            f"<div style='margin-top: 24px; padding: 12px; background-color: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 6px; font-size: 11px; color: #475569;'>"
            f"  <b style='color: #1e293b;'>System Efficiency Metadata:</b><br>"
            f"  &bull; Mail Received: {html.escape(system_efficiency['received_time'])}<br>"
            f"  &bull; Response Generated: {html.escape(system_efficiency['generated_time'])}<br>"
            f"  &bull; Processing Latency: {system_efficiency['latency']:.2f} seconds"
            f"</div>"
        )
    html_lines.append("</div></body></html>")
    html_text = "\n".join(html_lines)

    return (plain_text, html_text), grand_total


def build_empty_reply_body(customer_name, logo_cid=None, tenant_config=None, system_efficiency=None):
    """Formats a reply body for when no order items could be parsed (both Plain Text and HTML)."""
    tenant_id = tenant_config.get("id", "default") if tenant_config else "default"
    from src.database_sqlite import get_setting
    exec_name = get_setting("exec_name", tenant_config.get("sales_executive_name", SALES_EXECUTIVE_NAME) if tenant_config else SALES_EXECUTIVE_NAME, tenant_id)
    exec_title = get_setting("exec_title", tenant_config.get("sales_executive_title", SALES_EXECUTIVE_TITLE) if tenant_config else SALES_EXECUTIVE_TITLE, tenant_id)
    bus_name = get_setting("business_name", tenant_config.get("business_name", BUSINESS_NAME) if tenant_config else BUSINESS_NAME, tenant_id)
    exec_phone = get_setting("exec_phone", tenant_config.get("sales_executive_phone", SALES_EXECUTIVE_PHONE) if tenant_config else SALES_EXECUTIVE_PHONE, tenant_id)
    exec_email = get_setting("exec_email", tenant_config.get("sales_executive_email", SALES_EXECUTIVE_EMAIL) if tenant_config else SALES_EXECUTIVE_EMAIL, tenant_id)

    # Plain text
    body = [
        f"Dear {customer_name},\n",
        "Thank you for getting in touch with us.\n",
        "We went through your message but had a bit of trouble identifying the specific items and quantities you need. Could you send us a clearer list so we can get you the pricing quickly?\n",
        "For example:",
        "  - 10 x Brass Threaded Elbow Fitting 1/2 Inch",
        "  - 5 x PTFE Teflon Seal Tape 12mm\n",
        "Once we have the details, we'll get back to you with the quote right away.",
        "\nWarm regards,",
        exec_name,
        f"{exec_title} | {bus_name}",
        exec_phone
    ]
    if system_efficiency:
        body.append("\n" + "=" * 40)
        body.append("System Efficiency Metadata:")
        body.append(f"- Mail Received: {system_efficiency['received_time']}")
        body.append(f"- Response Generated: {system_efficiency['generated_time']}")
        body.append(f"- Processing Latency: {system_efficiency['latency']:.2f} seconds")
        body.append("=" * 40)
        
    plain_text = "\n".join(body)
    
    # HTML
    html_lines = [
        "<html><head><style>",
        "body { font-family: Arial, 'Helvetica Neue', sans-serif; color: #334155; line-height: 1.6; margin: 0; padding: 24px; font-size: 14px; }",
        ".footer { margin-top: 36px; border-top: 1px solid #e2e8f0; padding-top: 18px; color: #64748b; font-size: 13px; }",
        ".footer b { color: #1e293b; }",
        "</style></head><body>",
        f"<p>Dear {html.escape(customer_name)},</p>",
        "<p>Thank you for getting in touch with us.</p>",
        "<p>We went through your message but had a bit of trouble identifying the specific items and quantities you need. Could you send us a clearer list so we can get you the pricing quickly?</p>",
        "<p>For example:</p>",
        "<ul style='color:#334155;'>",
        "<li>10 x Brass Threaded Elbow Fitting 1/2 Inch</li>",
        "<li>5 x PTFE Teflon Seal Tape 12mm</li>",
        "</ul>",
        "<p>Once we have the details, we&#39;ll get back to you with the quote right away.</p>",
        "<div class='footer'>",
    ]
    if logo_cid:
        html_lines.append(f"<img src='cid:{logo_cid}' alt='{bus_name}' style='max-height: 50px; margin-bottom: 12px; display:block;'>")
    html_lines.append(f"Warm regards,<br><b>{exec_name}</b><br>{exec_title} &nbsp;|&nbsp; {bus_name}<br><span style='color:#94a3b8;'>{exec_phone} &nbsp;&bull;&nbsp; {exec_email}</span>")
    
    if system_efficiency:
        html_lines.append(
            f"<div style='margin-top: 24px; padding: 12px; background-color: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 6px; font-size: 11px; color: #475569;'>"
            f"  <b style='color: #1e293b;'>System Efficiency Metadata:</b><br>"
            f"  &bull; Mail Received: {html.escape(system_efficiency['received_time'])}<br>"
            f"  &bull; Response Generated: {html.escape(system_efficiency['generated_time'])}<br>"
            f"  &bull; Processing Latency: {system_efficiency['latency']:.2f} seconds"
            f"</div>"
        )
        
    html_lines.append("</div></body></html>")
    
    html_text = "\n".join(html_lines)
    return plain_text, html_text

def is_negotiation_msg(text):
    """Checks if the text contains negotiation-related keywords with word boundaries and phrase detection."""
    if not text:
        return False
    text_lower = text.lower()
    
    # 1. Keywords that must match as full words
    word_keywords = [
        r"discounts?", r"cheaper", r"reductions?", r"reduce", 
        r"negotiat\w*", r"deals?", r"concessions?"
    ]
    # Check if any full word keyword matches
    for kw in word_keywords:
        if re.search(r'\b' + kw + r'\b', text_lower):
            return True
            
    # 2. Percentage match: e.g. "10%" or "10 %" or "10 percent"
    if re.search(r'\b\d+(?:\.\d+)?\s*(?:%|percent)\b', text_lower):
        return True
        
    # 3. Specific price reduction phrases
    phrases = [
        r"better\s+(?:price|rate|deal|offer)",
        r"lower\s+(?:price|rate|cost|amount|total)",
        r"less\s+(?:price|money|cost|amount)",
        r"give\s+(?:me\s+)?(?:a\s+)?(?:better\s+)?(?:price|discount|deal)",
        r"make\s+(?:it\s+)?cheaper",
        r"price\s+(?:is\s+)?too\s+high"
    ]
    for phrase in phrases:
        if re.search(phrase, text_lower):
            return True
            
    return False

def is_human_request(text):
    """Checks if the text contains requests for human related action."""
    if not text:
        return False
    text_clean = strip_email_history(text)
    keywords = [
        "human", "salesperson", "sales executive", "representative", "agent", 
        "call me", "manager", "support staff", "operator", "talk to someone", 
        "person", "contact me back", "phone me", "speak to", "real assistant"
    ]
    text_lower = text_clean.lower()
    return any(re.search(r'\b' + re.escape(kw) + r'\b', text_lower) for kw in keywords)


def extract_requested_discount(text):
    """Extracts requested discount percentage from message body using regex patterns."""
    # Pattern 1: 10%
    pct_match = re.search(r'(\d+(?:\.\d+)?)\s*%', text)
    if pct_match:
        return float(pct_match.group(1))
    
    # Pattern 2: 10 percent
    pct_match_word = re.search(r'(\d+(?:\.\d+)?)\s*percent', text, re.IGNORECASE)
    if pct_match_word:
        return float(pct_match_word.group(1))
        
    # Pattern 3: discount of 10
    off_match = re.search(r'(?:discount|off|reduce)\s+(?:of\s+)?(\d+(?:\.\d+)?)', text, re.IGNORECASE)
    if off_match:
        return float(off_match.group(1))
        
    # Default to 5% if unspecified
    return 5.0

def extract_global_quantity_override(text):
    """
    Looks for phrases like 'provide me with 15 quantities respectively' or 'quote 15 of each' 
    and returns the extracted override quantity as a float/int.
    """
    text_clean = text.lower().strip()
    
    # Pattern 1: 'provide me with 15 quantities respectively'
    # Pattern 2: 'quote 15 of each' / 'send 15 of each'
    # Pattern 3: '15 quantities respectively'
    pattern = r'\b(?:provide|quote|send|need|want|give|deliver|with)\s+(?:me\s+)?(?:with\s+)?(\d+(?:\.\d+)?)\s*(?:quantities|qty|pcs|items|units|rolls|joints|count)?\s*(?:respectively|of\s+each|each)\b'
    match = re.search(pattern, text_clean)
    if match:
        try:
            val = float(match.group(1))
            return int(val) if val.is_integer() else val
        except ValueError:
            pass
            
    # Pattern 4: '\b(\d+)\s+(?:quantities|qty|pcs|units|items|rolls|joints|count)\s*respectively\b'
    pattern_alt = r'\b(\d+(?:\.\d+)?)\s*(?:quantities|qty|pcs|units|items|rolls|joints|count)?\s*respectively\b'
    match_alt = re.search(pattern_alt, text_clean)
    if match_alt:
        try:
            val = float(match_alt.group(1))
            return int(val) if val.is_integer() else val
        except ValueError:
            pass
            
    return None

class EmailResponseTuple(tuple):
    def __new__(cls, reply_subject, reply_body, pdf_path, status, invoice_id=None):
        obj = super().__new__(cls, (reply_subject, reply_body, pdf_path, status))
        obj.invoice_id = invoice_id
        return obj

def process_incoming_email(sender, subject, body, catalog, crm_path, mode, project_root, tenant_id=None, prior_thread_context=None, skip_initial_customer_log=False):
    """
    Main ingestion business logic. Parses the email body, matches SKUs, 
    manages CRM discounts, handles price negotiations, and returns the response.
    Returns: (reply_subject, reply_body_tuple, pdf_path, status)
    
    If prior_thread_context is provided (a plain-text conversation summary from
    prior chat_logs), it is prepended to the body before AI processing so the
    engine can answer follow-up questions with full conversation awareness.
    """
    import time
    from datetime import datetime, timedelta
    start_time = time.time()
    # Mail received time is recorded when system starts ingestion/processing
    now_ist = datetime.utcnow() + timedelta(hours=5, minutes=30)
    received_time_str = now_ist.strftime('%Y-%m-%d %H:%M:%S IST')

    from src.tenants import get_tenant_config
    tenant_config = get_tenant_config(tenant_id)
    
    # Get Gemini client for query batching / vector matching
    client = _get_gemini_client()
    
    # Dynamically resolve CRM path from tenant config if configured
    if tenant_config and tenant_config.get("crm_json"):
        tenant_crm = tenant_config.get("crm_json")
        if not os.path.isabs(tenant_crm):
            crm_path = os.path.join(project_root, tenant_crm)
        else:
            crm_path = tenant_crm
        if not os.path.exists(crm_path):
            crm_path = os.path.join(project_root, "data", "crm_customers.json")

    body_clean = strip_email_history(body)
    latest_customer_message = body_clean

    # ── Thread Context Injection ──────────────────────────────────────────────
    # If the caller provided prior conversation history (for customer replies),
    # prepend it so the AI engine has full context when processing the new message.
    if prior_thread_context and prior_thread_context.strip():
        body_clean = (
            "=== Prior Conversation History ===\n"
            + prior_thread_context.strip()
            + "\n=================================\n\n"
            + "[Customer's latest message:]\n"
            + body_clean
        )
    # ─────────────────────────────────────────────────────────────────────────
    
    # Determine if we should cap quantities based on stock availability

    # (Only cap if we have an attachment AND the customer did NOT write additional product info in the body)
    has_attachment = "[From attachment" in latest_customer_message
    cap_by_stock = False
    override_qty = None
    
    if has_attachment:
        cap_by_stock = True
        # Split latest_customer_message to separate original text from attachment text
        parts = latest_customer_message.split("[From attachment")
        email_text_only = parts[0].strip()
        
        override_qty = extract_global_quantity_override(email_text_only)
        
        # Check if the customer gave additional info based on product in the email text
        body_lines = run_scenario_free(email_text_only, catalog, gemini_client=client)
        body_has_products = any(l["matched_sku_id"] != "UNKNOWN" for l in body_lines)
        
        if body_has_products or override_qty is not None:
            cap_by_stock = False
    
    import email.utils
    display_name, email_addr = email.utils.parseaddr(sender)
    if not email_addr:
        email_addr = sender
    
    # Check if this is a thread reply to an existing quotation
    # First check subject line, then scan body (customers often reply without explicit QTN in subject)
    quote_id_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', subject, re.IGNORECASE)
    if not quote_id_match:
        # Scan body for QTN reference (from thread history)
        body_id_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', body[:3000], re.IGNORECASE)
        if body_id_match:
            quote_id_match = body_id_match
    if not quote_id_match:
        # QTN-XXXXX pattern in subject or body
        subj_qtn_match = re.search(r'(QTN-[A-Z0-9\-]+)', subject, re.IGNORECASE)
        if subj_qtn_match:
            quote_id_match = subj_qtn_match
        else:
            body_qtn_match = re.search(r'(QTN-[A-Z0-9\-]+)', body[:3000], re.IGNORECASE)
            if body_qtn_match:
                quote_id_match = body_qtn_match
    
    meta_path = None
    meta = None
    existing_invoice_id = None
    
    if quote_id_match:
        existing_invoice_id = quote_id_match.group(1).upper()
        meta_filename = f"Quote_{existing_invoice_id}_meta.json"
        
        # Meta files might be stored in static/quotes, static/quotes/<tenant_id>, mock_outbox, or mock_outbox/<tenant_id>
        meta_paths = [
            os.path.join(project_root, "static", "quotes", meta_filename),
        ]
        if tenant_id and tenant_id != "default":
            meta_paths.append(os.path.join(project_root, "static", "quotes", tenant_id, meta_filename))
            meta_paths.append(os.path.join(project_root, "mock_outbox", tenant_id, meta_filename))
        meta_paths.append(os.path.join(project_root, "mock_outbox", meta_filename))
        
        for p in meta_paths:
            if os.path.exists(p):
                meta_path = p
                try:
                    with open(p, 'r', encoding='utf-8') as f:
                        meta = json.load(f)
                    break
                except Exception:
                    pass
        
        # Bug 2 fix: If meta JSON file is missing, rebuild meta from SQLite database
        if not meta and existing_invoice_id:
            try:
                from src.database_sqlite import get_connection
                _mc = get_connection(tenant_id)
                _qrow = _mc.execute(
                    "SELECT customer_name, customer_email, discount_pct, subtotal FROM quotations WHERE invoice_id = ? LIMIT 1",
                    (existing_invoice_id,)
                ).fetchone()
                if _qrow:
                    _items = _mc.execute(
                        "SELECT sku_id, sku_name, quantity, unit_price FROM quotation_items WHERE invoice_id = ?",
                        (existing_invoice_id,)
                    ).fetchall()
                    _mc.close()
                    meta = {
                        "invoice_id": existing_invoice_id,
                        "customer_name": _qrow["customer_name"],
                        "customer_email": _qrow["customer_email"],
                        "customer_phone": "—",
                        "discount_pct": _qrow["discount_pct"] or 0.0,
                        "matched_lines": [
                            {"matched_sku_id": r["sku_id"], "matched_sku_name": r["sku_name"],
                             "quantity": r["quantity"], "unit_price": r["unit_price"],
                             "confidence": 100.0, "deficit": 0}
                            for r in _items
                        ],
                        "chat_history": []
                    }
                    # Use a temp path in static/quotes for writing back
                    if tenant_id and tenant_id != "default":
                        _mdir = os.path.join(project_root, "static", "quotes", tenant_id)
                    else:
                        _mdir = os.path.join(project_root, "static", "quotes")
                    os.makedirs(_mdir, exist_ok=True)
                    meta_path = os.path.join(_mdir, meta_filename)
                    print(f"[process_incoming_email] Rebuilt meta from DB for {existing_invoice_id} (meta file was missing).")
                else:
                    _mc.close()
            except Exception as _me:
                print(f"[process_incoming_email] DB meta rebuild failed for {existing_invoice_id}: {_me}")

    # Process conversational thread replies if we have the quotation metadata
    if meta and existing_invoice_id:
        # Run ingestion first to see if they are modifying/adding products
        matched_lines = _run_ingestion(latest_customer_message, catalog, client, tenant_id)
        has_valid_matches = matched_lines and any(
            line['matched_sku_id'] != "UNKNOWN" and line['confidence'] >= 80.0 
            for line in matched_lines
        )
        
        # 1. Check if this is a negotiation request (takes priority ONLY if they are not adding/modifying products)
        if is_negotiation_msg(latest_customer_message) and not has_valid_matches:
            requested_discount = extract_requested_discount(latest_customer_message)
            chat_history = meta.get("chat_history", [])
            chat_history.append({"sender": "customer", "text": latest_customer_message})
            
            # Setup GenAI client for live mode if API key is valid
            api_key = os.environ.get("GEMINI_API_KEY")
            client = None
            is_live = False
            if api_key and api_key.strip() and not api_key.startswith("your_"):
                try:
                    from google import genai
                    client = genai.Client(api_key=api_key)
                    is_live = True
                except Exception:
                    pass
            
            from src.negotiator import run_negotiation_step
            neg_result = run_negotiation_step(
                customer_message=latest_customer_message,
                requested_discount=requested_discount,
                chat_history=chat_history,
                is_live=is_live,
                client=client
            )
            
            status = neg_result.get("status", "NEGOTIATING")
            approved_discount = neg_result.get("approved_discount", 0.0)
            reply_text = neg_result.get("reply", "")
            
            chat_history.append({"sender": "ai", "text": reply_text})
            meta["chat_history"] = chat_history
            
            pdf_out_path = None
            if status == "APPROVED":
                new_discount_pct = approved_discount / 100.0
                meta["discount_pct"] = new_discount_pct
                
                if mode == "mock":
                    if tenant_id and tenant_id != "default":
                        pdf_out_path = os.path.join(project_root, "mock_outbox", tenant_id, f"Quote_{existing_invoice_id}.pdf")
                    else:
                        pdf_out_path = os.path.join(project_root, "mock_outbox", f"Quote_{existing_invoice_id}.pdf")
                else:
                    if tenant_id and tenant_id != "default":
                        pdf_out_path = os.path.join(project_root, "static", "quotes", tenant_id, f"Quote_{existing_invoice_id}.pdf")
                    else:
                        pdf_out_path = os.path.join(project_root, "static", "quotes", f"Quote_{existing_invoice_id}.pdf")
                
                any_quoted = any(line['quantity'] > 0 for line in meta["matched_lines"] if line.get('matched_sku_id') != 'UNKNOWN')
                if any_quoted:
                    generate_pdf_quotation(
                        matched_lines=meta["matched_lines"],
                        discount_pct=new_discount_pct,
                        customer_name=meta["customer_name"],
                        invoice_id=existing_invoice_id,
                        output_path=pdf_out_path,
                        catalog=catalog,
                        customer_phone=meta.get("customer_phone", "—"),
                        customer_email=meta.get("customer_email"),
                        upi_id=tenant_config.get("upi_id"),
                        upi_name=tenant_config.get("upi_name"),
                        logo_path=tenant_config.get("company_logo_path"),
                        business_name=tenant_config.get("business_name")
                    )
                else:
                    pdf_out_path = None
                
                duration = time.time() - start_time
                now_ist = datetime.utcnow() + timedelta(hours=5, minutes=30)
                generated_time_str = now_ist.strftime('%Y-%m-%d %H:%M:%S IST')
                system_efficiency = {
                    "received_time": received_time_str,
                    "generated_time": generated_time_str,
                    "latency": duration
                }

                reply_body, grand_total = build_email_reply_body(
                    matched_lines=meta["matched_lines"],
                    discount_pct=new_discount_pct,
                    customer_name=meta["customer_name"],
                    invoice_id=existing_invoice_id,
                    logo_cid="company_logo",
                    tenant_config=tenant_config,
                    customer_email=meta.get("customer_email"),
                    customer_phone=meta.get("customer_phone"),
                    system_efficiency=system_efficiency
                )
                
                # Log to SQLite
                try:
                    from src.database_sqlite import update_quotation_status, log_chat_msg
                    update_quotation_status(existing_invoice_id, "NEGOTIATION_APPROVED", new_discount_pct, tenant_id=tenant_id)
                    log_chat_msg(existing_invoice_id, "customer", latest_customer_message, tenant_id=tenant_id)
                    log_chat_msg(existing_invoice_id, "ai", reply_text, tenant_id=tenant_id)
                except Exception as e:
                    print(f"[Warning] SQLite logging failed: {e}")
                
                # Combine AI reply and table breakdown
                full_reply_plain = f"{reply_text}\n\n{reply_body[0]}"
                full_reply_html = f"<html><body><p>{reply_text.replace('\n', '<br>')}</p>{reply_body[1].replace('<html><body>', '').replace('</body></html>', '')}</body></html>"
                reply_payload = (full_reply_plain, full_reply_html)
            else:
                duration = time.time() - start_time
                now_ist = datetime.utcnow() + timedelta(hours=5, minutes=30)
                generated_time_str = now_ist.strftime('%Y-%m-%d %H:%M:%S IST')
                
                efficiency_plain = (
                    f"\n\n{'=' * 40}\n"
                    f"System Efficiency Metadata:\n"
                    f"- Mail Received: {received_time_str}\n"
                    f"- Response Generated: {generated_time_str}\n"
                    f"- Processing Latency: {duration:.2f} seconds\n"
                    f"{'=' * 40}"
                )
                efficiency_html = (
                    f"<div style='margin-top: 24px; padding: 12px; background-color: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 6px; font-size: 11px; color: #475569;'>"
                    f"  <b style='color: #1e293b;'>System Efficiency Metadata:</b><br>"
                    f"  &bull; Mail Received: {received_time_str}<br>"
                    f"  &bull; Response Generated: {generated_time_str}<br>"
                    f"  &bull; Processing Latency: {duration:.2f} seconds"
                    f"</div>"
                )
                
                full_reply_plain = reply_text + efficiency_plain
                full_reply_html = f"<html><body><p>{reply_text.replace('\n', '<br>')}</p>{efficiency_html}</body></html>"
                reply_payload = (full_reply_plain, full_reply_html)
                
                # Log to SQLite
                try:
                    from src.database_sqlite import update_quotation_status, log_chat_msg
                    db_status = "PENDING_REVIEW" if status == "PENDING_REVIEW" else f"NEGOTIATION_{status}"
                    update_quotation_status(existing_invoice_id, db_status, tenant_id=tenant_id)
                    log_chat_msg(existing_invoice_id, "customer", latest_customer_message, tenant_id=tenant_id)
                    log_chat_msg(existing_invoice_id, "ai", reply_text, tenant_id=tenant_id)
                except Exception as e:
                    print(f"[Warning] SQLite logging failed: {e}")
                    
                # Write back meta updates for intermediate negotiation/escalation turns
                try:
                    with open(meta_path, 'w', encoding='utf-8') as f:
                        json.dump(meta, f, indent=2)
                except Exception:
                    pass
            
            reply_subject = clean_reply_subject(subject, invoice_id=existing_invoice_id)
            db_status = "PENDING_REVIEW" if status == "PENDING_REVIEW" else f"NEGOTIATION_{status}"
            return EmailResponseTuple(reply_subject, reply_payload, pdf_out_path, db_status, existing_invoice_id)
            
        # 2. Check if the customer is modifying the quotation items (high-confidence items only!)
        # We already computed matched_lines above!

        # ── Deduplicate: merge rows with the same SKU ID, summing quantities ──
        _dedup_map2 = {}
        _dedup_order2 = []
        for line in matched_lines:
            sku_id = line.get('matched_sku_id', 'UNKNOWN')
            if sku_id == 'UNKNOWN':
                _dedup_order2.append(line)
                continue
            if sku_id in _dedup_map2:
                _dedup_map2[sku_id]['quantity'] += line['quantity']
            else:
                _dedup_map2[sku_id] = dict(line)
                _dedup_order2.append(_dedup_map2[sku_id])
        matched_lines = _dedup_order2

        if override_qty is not None:
            for line in matched_lines:
                if line['matched_sku_id'] != "UNKNOWN":
                    line['quantity'] = override_qty

        # Extract customer details and check for new phone number
        customer_name = meta.get("customer_name", "Walk-in Retail Client")
        customer_phone = meta.get("customer_phone", "—")
        new_phone = extract_phone_number(body_clean)
        if new_phone:
            customer_phone = new_phone
            meta["customer_phone"] = customer_phone

        adjust_quantities_by_stock(
            matched_lines,
            catalog,
            cap_by_stock=cap_by_stock,
            invoice_id=existing_invoice_id,
            customer_name=customer_name,
            customer_email=email_addr,
            customer_phone=customer_phone,
            tenant_id=tenant_id
        )
        has_valid_matches = matched_lines and any(
            line['matched_sku_id'] != "UNKNOWN" and line['confidence'] >= 80.0 
            for line in matched_lines
        )
        
        if has_valid_matches:
            # Re-generate the quote with the new items under the same ID
            discount_pct = meta.get("discount_pct", 0.0)
            
            if mode == "mock":
                if tenant_id and tenant_id != "default":
                    pdf_out_path = os.path.join(project_root, "mock_outbox", tenant_id, f"Quote_{existing_invoice_id}.pdf")
                else:
                    pdf_out_path = os.path.join(project_root, "mock_outbox", f"Quote_{existing_invoice_id}.pdf")
            else:
                if tenant_id and tenant_id != "default":
                    pdf_out_path = os.path.join(project_root, "static", "quotes", tenant_id, f"Quote_{existing_invoice_id}.pdf")
                else:
                    pdf_out_path = os.path.join(project_root, "static", "quotes", f"Quote_{existing_invoice_id}.pdf")
            
            any_quoted = any(line['quantity'] > 0 for line in matched_lines if line.get('matched_sku_id') != 'UNKNOWN')
            if any_quoted:
                generate_pdf_quotation(
                    matched_lines=matched_lines,
                    discount_pct=discount_pct,
                    customer_name=customer_name,
                    invoice_id=existing_invoice_id,
                    output_path=pdf_out_path,
                    catalog=catalog,
                    customer_phone=customer_phone,
                    customer_email=email_addr,
                    upi_id=tenant_config.get("upi_id"),
                    upi_name=tenant_config.get("upi_name"),
                    logo_path=tenant_config.get("company_logo_path"),
                    business_name=tenant_config.get("business_name")
                )
            else:
                pdf_out_path = None
                
            duration = time.time() - start_time
            now_ist = datetime.utcnow() + timedelta(hours=5, minutes=30)
            generated_time_str = now_ist.strftime('%Y-%m-%d %H:%M:%S IST')
            system_efficiency = {
                "received_time": received_time_str,
                "generated_time": generated_time_str,
                "latency": duration
            }

            reply_body, grand_total = build_email_reply_body(
                matched_lines=matched_lines,
                discount_pct=discount_pct,
                customer_name=customer_name,
                invoice_id=existing_invoice_id,
                logo_cid="company_logo",
                tenant_config=tenant_config,
                customer_email=email_addr,
                customer_phone=customer_phone,
                system_efficiency=system_efficiency
            )
            
            # Log to SQLite
            try:
                from src.database_sqlite import log_quotation, log_quotation_item
                raw_subtotal = sum(i["quantity"] * i["unit_price"] for i in matched_lines if i["matched_sku_id"] != "UNKNOWN")
                discount_amt = raw_subtotal * discount_pct
                net_subtotal = raw_subtotal - discount_amt
                tax_amt = net_subtotal * 0.18
                
                log_quotation(
                    invoice_id=existing_invoice_id,
                    customer_name=customer_name,
                    customer_email=email_addr,
                    customer_phone=customer_phone,
                    subtotal=raw_subtotal,
                    discount_pct=discount_pct,
                    tax_amt=tax_amt,
                    grand_total=grand_total,
                    status="QUOTE_UPDATED",
                    tenant_id=tenant_id
                )
                for item in matched_lines:
                    if item["matched_sku_id"] != "UNKNOWN":
                        log_quotation_item(
                            invoice_id=existing_invoice_id,
                            sku_id=item["matched_sku_id"],
                            sku_name=item["matched_sku_name"],
                            quantity=item["quantity"],
                            unit_price=item["unit_price"],
                            line_total=item["quantity"] * item["unit_price"],
                            tenant_id=tenant_id
                        )
                # Log the customer email and bot reply for the View Request modal
                from src.database_sqlite import log_chat_msg
                try:
                    if not skip_initial_customer_log:
                        log_chat_msg(existing_invoice_id, "CUSTOMER", latest_customer_message, tenant_id=tenant_id)
                    plain_reply = reply_body[0] if isinstance(reply_body, tuple) else str(reply_body)
                    log_chat_msg(existing_invoice_id, "BOT", plain_reply, tenant_id=tenant_id)
                except Exception as _ce:
                    print(f"[Warning] chat_log insert failed: {_ce}")
            except Exception as e:
                print(f"[Warning] SQLite logging failed: {e}")
                
            reply_subject = clean_reply_subject(subject, invoice_id=existing_invoice_id)
            return EmailResponseTuple(reply_subject, reply_body, pdf_out_path, "QUOTE_UPDATED", existing_invoice_id)

        else:
            # Handle general conversational enquiry
            chat_history = meta.get("chat_history", [])
            chat_history.append({"sender": "customer", "text": latest_customer_message})
            
            api_key = os.environ.get("GEMINI_API_KEY")
            client = None
            is_live = False
            if api_key and api_key.strip() and not api_key.startswith("your_"):
                try:
                    from google import genai
                    client = genai.Client(api_key=api_key)
                    is_live = True
                except Exception:
                    pass
            
            if is_live and client:
                history_formatted = "\n".join([f"{msg['sender']}: {msg['text']}" for msg in chat_history[:-1]])
                prompt = f"""
                You are a sales assistant representing {tenant_config.get('business_name', 'Trofeo Hardware')}.
                A customer is replying to their existing quotation #{existing_invoice_id}.
                
                Here is the quotation metadata:
                - Customer Name: {meta['customer_name']}
                - Current Discount Applied: {meta['discount_pct']*100}%
                - Items Quoted: {json.dumps(meta['matched_lines'])}
                
                Here is the conversation history:
                {history_formatted}
                
                Customer's latest reply: "{latest_customer_message}"
                
                Write a concise, polite, and professional response. If they are confirming the order, thank them and let them know the team will prepare the final invoice for payment. If they are asking a product question, answer it if you can based on the items, or politely state you are forwarding this to a salesperson.
                """
                try:
                    response = client.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=prompt
                    )
                    reply_text = response.text.strip()
                except Exception:
                    reply_text = f"Thank you for your message. We have received your query regarding Quotation #{existing_invoice_id} and escalated it to our sales desk. A representative will contact you shortly."
            else:
                reply_text = f"Thank you for your message. We have received your query regarding Quotation #{existing_invoice_id} and escalated it to our sales desk. A representative will contact you shortly."
                
            chat_history.append({"sender": "ai", "text": reply_text})
            meta["chat_history"] = chat_history
            
            try:
                from src.database_sqlite import update_quotation_status, log_chat_msg
                update_quotation_status(existing_invoice_id, "CONVERSATIONAL_REPLY", tenant_id=tenant_id)
                # Bug 6 fix: Log bot reply to chat_logs for timeline visibility
                log_chat_msg(existing_invoice_id, "BOT", reply_text, tenant_id=tenant_id)
            except Exception as e:
                print(f"[Warning] Failed to update quotation status to CONVERSATIONAL_REPLY: {e}")
                
            try:
                with open(meta_path, 'w', encoding='utf-8') as f:
                    json.dump(meta, f, indent=2)
            except Exception:
                pass
                
            reply_subject = clean_reply_subject(subject, invoice_id=existing_invoice_id)
            reply_payload = (reply_text, f"<html><body><p>{reply_text.replace('\n', '<br>')}</p></body></html>")
            return EmailResponseTuple(reply_subject, reply_payload, None, "CONVERSATIONAL_REPLY", existing_invoice_id)

    # Default logic: Brand new quotation request
    discount_pct, customer_name, customer_phone = get_crm_discount(email_addr, crm_path)
    if customer_name == "Walk-in Retail Client":
        if display_name:
            customer_name = display_name
        else:
            customer_name = email_addr.split('@')[0]
            
    # Extract phone number if present in email body
    extracted_phone = extract_phone_number(body_clean)
    if extracted_phone:
        customer_phone = extracted_phone
            
    from src.database_sqlite import generate_next_invoice_id
    invoice_id = generate_next_invoice_id(tenant_id=tenant_id)

    matched_lines = _run_ingestion(body_clean, catalog, client, tenant_id)

    # ── Deduplicate: merge rows with the same SKU ID, summing quantities ──
    # This handles cases where the parser returns the same item multiple times
    # (e.g. from copy-pasted tables where a SKU appears in both header and data rows)
    _dedup_map = {}
    _dedup_order = []
    for line in matched_lines:
        sku_id = line.get('matched_sku_id', 'UNKNOWN')
        if sku_id == 'UNKNOWN':
            _dedup_order.append(line)
            continue
        if sku_id in _dedup_map:
            # Merge: add quantities, keep other fields from first occurrence
            _dedup_map[sku_id]['quantity'] += line['quantity']
        else:
            _dedup_map[sku_id] = dict(line)
            _dedup_order.append(_dedup_map[sku_id])
    matched_lines = _dedup_order
    print(f"[Dedup] {len(matched_lines)} unique item(s) after deduplication.")

    if override_qty is not None:
        for line in matched_lines:
            if line['matched_sku_id'] != "UNKNOWN":
                line['quantity'] = override_qty

    deficit_lines = adjust_quantities_by_stock(
        matched_lines, 
        catalog, 
        cap_by_stock=cap_by_stock,
        invoice_id=invoice_id,
        customer_name=customer_name,
        customer_email=email_addr,
        customer_phone=customer_phone,
        tenant_id=tenant_id
    )
    
    has_valid_matches = matched_lines and any(
        line['matched_sku_id'] != "UNKNOWN" and line['confidence'] >= 80.0 
        for line in matched_lines
    )
    
    if has_valid_matches:
        any_quoted = any(line['quantity'] > 0 for line in matched_lines if line.get('matched_sku_id') != 'UNKNOWN')
        pdf_out_path = None
        if any_quoted:
            if mode == "mock":
                if tenant_id and tenant_id != "default":
                    pdf_out_path = os.path.join(project_root, "mock_outbox", tenant_id, f"Quote_{invoice_id}.pdf")
                else:
                    pdf_out_path = os.path.join(project_root, "mock_outbox", f"Quote_{invoice_id}.pdf")
            else:
                if tenant_id and tenant_id != "default":
                    pdf_out_path = os.path.join(project_root, "static", "quotes", tenant_id, f"Quote_{invoice_id}.pdf")
                else:
                    pdf_out_path = os.path.join(project_root, "static", "quotes", f"Quote_{invoice_id}.pdf")
                
            generate_pdf_quotation(
                matched_lines=matched_lines,
                discount_pct=discount_pct,
                customer_name=customer_name,
                invoice_id=invoice_id,
                output_path=pdf_out_path,
                catalog=catalog,
                customer_phone=customer_phone,
                upi_id=tenant_config.get("upi_id"),
                upi_name=tenant_config.get("upi_name"),
                logo_path=tenant_config.get("company_logo_path"),
                business_name=tenant_config.get("business_name"),
                customer_email=email_addr
            )
        duration = time.time() - start_time
        now_ist = datetime.utcnow() + timedelta(hours=5, minutes=30)
        generated_time_str = now_ist.strftime('%Y-%m-%d %H:%M:%S IST')
        system_efficiency = {
            "received_time": received_time_str,
            "generated_time": generated_time_str,
            "latency": duration
        }

        reply_body, grand_total = build_email_reply_body(
            matched_lines=matched_lines,
            discount_pct=discount_pct,
            customer_name=customer_name,
            invoice_id=invoice_id,
            logo_cid="company_logo",
            tenant_config=tenant_config,
            customer_email=email_addr,
            customer_phone=customer_phone,
            system_efficiency=system_efficiency
        )
        
        # Log to SQLite
        try:
            from src.database_sqlite import log_quotation, log_quotation_item
            raw_subtotal = sum(i["quantity"] * i["unit_price"] for i in matched_lines if i["matched_sku_id"] != "UNKNOWN")
            discount_amt = raw_subtotal * discount_pct
            net_subtotal = raw_subtotal - discount_amt
            tax_amt = net_subtotal * 0.18
            
            log_quotation(
                invoice_id=invoice_id,
                customer_name=customer_name,
                customer_email=email_addr,
                customer_phone=customer_phone,
                subtotal=raw_subtotal,
                discount_pct=discount_pct,
                tax_amt=tax_amt,
                grand_total=grand_total,
                status="QUOTE_GENERATED",
                tenant_id=tenant_id
            )
            for item in matched_lines:
                if item["matched_sku_id"] != "UNKNOWN":
                    log_quotation_item(
                        invoice_id=invoice_id,
                        sku_id=item["matched_sku_id"],
                        sku_name=item["matched_sku_name"],
                        quantity=item["quantity"],
                        unit_price=item["unit_price"],
                        line_total=item["quantity"] * item["unit_price"],
                        tenant_id=tenant_id
                    )
                # Log the customer email and bot reply for the View Request modal
                from src.database_sqlite import log_chat_msg
                try:
                    log_chat_msg(invoice_id, "CUSTOMER", body_clean, tenant_id=tenant_id)
                    plain_reply = reply_body[0] if isinstance(reply_body, tuple) else str(reply_body)
                    log_chat_msg(invoice_id, "BOT", plain_reply, tenant_id=tenant_id)
                except Exception as _ce:
                    print(f"[Warning] chat_log insert failed: {_ce}")
        except Exception as e:
            print(f"[Warning] SQLite logging failed: {e}")
            
        reply_subject = clean_reply_subject(subject, invoice_id=invoice_id)
        return EmailResponseTuple(reply_subject, reply_body, pdf_out_path, "QUOTE_GENERATED", invoice_id)
    else:
        duration = time.time() - start_time
        now_ist = datetime.utcnow() + timedelta(hours=5, minutes=30)
        generated_time_str = now_ist.strftime('%Y-%m-%d %H:%M:%S IST')
        system_efficiency = {
            "received_time": received_time_str,
            "generated_time": generated_time_str,
            "latency": duration
        }
        
        reply_body = build_empty_reply_body(
            customer_name,
            logo_cid="company_logo",
            tenant_config=tenant_config,
            system_efficiency=system_efficiency
        )
        reply_subject = clean_reply_subject(subject, is_unparsed=True)
        return EmailResponseTuple(reply_subject, reply_body, None, "UNPARSED_NOTICE", None)

def load_crm_emails(crm_path):
    """Loads email addresses of registered customers from CRM."""
    if not crm_path or not os.path.exists(crm_path):
        return set()
    try:
        with open(crm_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return {email_addr.lower().strip() for email_addr in data.keys()}
    except Exception as e:
        print(f"[Warning] Failed to load CRM customer emails: {e}")
        return set()


def extract_text_from_docx(docx_bytes):
    """
    Extracts text from DOCX bytes using standard zipfile and xml parsing.
    """
    import zipfile
    import xml.etree.ElementTree as ET
    from io import BytesIO
    try:
        with zipfile.ZipFile(BytesIO(docx_bytes)) as docx:
            xml_content = docx.read('word/document.xml')
            root = ET.fromstring(xml_content)
            namespaces = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            texts = []
            for elem in root.findall('.//w:t', namespaces):
                if elem.text:
                    texts.append(elem.text)
            return ' '.join(texts)
    except Exception as e:
        print(f"[Attachment] Error parsing docx: {e}")
        return ""


def extract_text_from_xlsx(xlsx_bytes):
    """
    Extracts text/cells from XLSX bytes using openpyxl.
    """
    import openpyxl
    from io import BytesIO
    try:
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
        lines = []
        for name in wb.sheetnames:
            sheet = wb[name]
            lines.append(f"Sheet: {name}")
            for row in sheet.iter_rows(values_only=True):
                if any(val is not None for val in row):
                    row_str = ", ".join(str(val).strip() for val in row if val is not None)
                    lines.append(row_str)
        return "\n".join(lines)
    except Exception as e:
        print(f"[Attachment] Error parsing xlsx: {e}")
        return ""


def extract_text_from_xls(xls_bytes):
    """
    Extracts text/cells from XLS bytes using xlrd.
    """
    import xlrd
    try:
        wb = xlrd.open_workbook(file_contents=xls_bytes)
        lines = []
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            lines.append(f"Sheet: {sheet.name}")
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                if any(val != "" and val is not None for val in row):
                    row_str = ", ".join(str(val).strip() for val in row if val != "" and val is not None)
                    lines.append(row_str)
        return "\n".join(lines)
    except Exception as e:
        print(f"[Attachment] Error parsing xls: {e}")
        return ""


def has_attachments(msg):
    """
    Quickly scans a parsed email message to check if it has ANY attachments.
    Does not download or decode them — just checks MIME metadata.
    """
    supported_extensions = (".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif",
                            ".tiff", ".bmp", ".docx", ".txt", ".rtf", ".html",
                            ".xml", ".csv", ".xls", ".xlsx", ".odt")
    supported_types = (
        "application/pdf", "image/jpeg", "image/jpg", "image/png",
        "image/webp", "image/gif", "image/tiff", "image/bmp",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/docx", "application/x-docx",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream"
    )
    for part in msg.walk():
        content_type = part.get_content_type()
        disposition = str(part.get("Content-Disposition", ""))
        filename = part.get_filename() or ""
        is_attach = "attachment" in disposition.lower() or filename != ""
        if is_attach and (content_type in supported_types or
                          content_type.startswith("image/") or
                          filename.lower().endswith(supported_extensions)):
            return True
    return False



def extract_text_via_local_ocr(payload, ext):
    """
    Saves the payload bytes to a temp image file and runs the local ocr.ps1 script on it.
    """
    import subprocess
    import tempfile
    import os
    
    if not ext:
        ext = ".png"
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    temp_dir = os.path.join(project_root, "data")
    os.makedirs(temp_dir, exist_ok=True)
    
    fd, temp_path = tempfile.mkstemp(suffix=ext, dir=temp_dir)
    try:
        with os.fdopen(fd, 'wb') as f:
            f.write(payload)
        
        ocr_script = os.path.join(project_root, "ocr.ps1")
        cmd = ["powershell", "-ExecutionPolicy", "Bypass", "-File", ocr_script, temp_path]
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=30)
        
        if result.returncode == 0:
            ocr_text = result.stdout.strip()
            print(f"[Local OCR] Extracted text ({len(ocr_text)} chars) from image.")
            return ocr_text
        else:
            print(f"[Local OCR Warning] PowerShell returned non-zero code: {result.stderr}")
            return ""
    except Exception as e:
        print(f"[Local OCR Warning] Failed to run local OCR: {e}")
        return ""
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass
def extract_text_from_pdf(pdf_bytes):
    """
    Extracts text from PDF bytes using pypdf.
    If no text is found (e.g. scanned PDF), attempts to extract images from pages
    and runs local OCR on each extracted image.
    """
    import pypdf
    from io import BytesIO
    import os
    try:
        reader = pypdf.PdfReader(BytesIO(pdf_bytes))
        text_parts = []
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text()
            if page_text and page_text.strip():
                text_parts.append(page_text)
            
            # Extract images and run OCR for scanned/image PDFs
            try:
                if hasattr(page, "images"):
                    for img_idx, image_file_object in enumerate(page.images):
                        img_data = image_file_object.data
                        img_name = image_file_object.name
                        _, img_ext = os.path.splitext(img_name)
                        if not img_ext:
                            img_ext = ".png"
                        ocr_text = extract_text_via_local_ocr(img_data, img_ext)
                        if ocr_text and ocr_text.strip():
                            text_parts.append(ocr_text)
            except Exception as img_err:
                print(f"[PDF OCR] Warning: Failed to extract images from page {i}: {img_err}")
                
        return "\n".join(text_parts)
    except Exception as e:
        print(f"[PDF OCR] Error parsing PDF: {e}")
        return ""


def extract_text_from_attachments(msg):
    """
    Scans all MIME parts of an email.message object for attachments (PDF, images, Word, text format)
    and extracts product item lists from each. Uses Gemini 2.5 Flash, falling back to local OCR for images.
    """
    api_key = os.environ.get("GEMINI_API_KEY", "")
    has_gemini = api_key and not api_key.startswith("your_") and api_key.strip() != ""

    gemini_native_mimes = {
        "application/pdf": ".pdf",
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "image/tiff": ".tiff",
        "image/bmp": ".bmp",
    }

    docx_mimes = {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/docx": ".docx",
        "application/x-docx": ".docx"
    }

    xlsx_mimes = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/xlsx": ".xlsx",
        "application/x-xlsx": ".xlsx"
    }

    xls_mimes = {
        "application/vnd.ms-excel": ".xls",
        "application/xls": ".xls",
        "application/x-xls": ".xls"
    }

    extracted_texts = []

    for part in msg.walk():
        content_type = part.get_content_type()
        disposition = part.get("Content-Disposition", "")
        filename = part.get_filename() or ""

        # Determine if it's an attachment we support
        is_supported = (
            content_type in gemini_native_mimes or
            content_type in docx_mimes or
            content_type in xlsx_mimes or
            content_type in xls_mimes or
            content_type.startswith("text/") or
            filename.lower().endswith((".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif", ".tiff", ".bmp", ".docx", ".txt", ".rtf", ".html", ".xml", ".csv", ".xlsx", ".xls"))
        )

        # Skip if not an attachment or not supported
        is_attachment = "attachment" in disposition.lower() or filename != ""
        if not is_attachment or not is_supported:
            continue

        payload = part.get_payload(decode=True)
        if not payload:
            continue

        # Extract file extension for logs
        ext = ""
        if filename:
            _, file_ext = os.path.splitext(filename)
            ext = file_ext.lower()
        else:
            if content_type in gemini_native_mimes:
                ext = gemini_native_mimes[content_type]
            elif content_type in docx_mimes:
                ext = ".docx"
            elif content_type in xlsx_mimes:
                ext = ".xlsx"
            elif content_type in xls_mimes:
                ext = ".xls"
            elif content_type.startswith("text/"):
                ext = ".txt"

        display_filename = filename or f"attachment{ext}"
        print(f"[Attachment] Detected attachment: '{display_filename}' ({content_type}). Processing text.")

        # Check if it is native image/PDF vs text/word format
        is_native_gemini = content_type in gemini_native_mimes or ext in [".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif", ".tiff", ".bmp"]
        is_image = ext in [".jpg", ".jpeg", ".png", ".webp", ".gif", ".tiff", ".bmp"]

        extracted = ""
        
        # 1. Try Gemini Multimodal / Text restructuring first if key is configured
        if has_gemini:
            try:
                from google import genai
                from google.genai import types as genai_types
                client = genai.Client(api_key=api_key)

                prompt = (
                    "You are a purchasing document analyser. Extract ALL product names, item descriptions, "
                    "SKU codes, part numbers, quantities, and specifications mentioned in this document. "
                    "Return them as a plain numbered list. Do NOT add any extra commentary or headings - "
                    "just list each line item or product request exactly as written. "
                    "If you cannot find any product items, reply with the single word: NONE."
                )

                # Retry up to 3 times for transient API errors (503, 429, 500)
                max_retries = 3
                for attempt in range(1, max_retries + 1):
                    try:
                        if is_native_gemini:
                            # Native Gemini Multimodal parsing
                            response = client.models.generate_content(
                                model="gemini-2.5-flash",
                                contents=[
                                    genai_types.Part.from_bytes(
                                        data=payload,
                                        mime_type=content_type if content_type in gemini_native_mimes else "application/pdf"
                                    ),
                                    prompt
                                 ]
                            )
                            extracted = response.text.strip() if response.text else ""
                        else:
                            # Local text extraction followed by Gemini text restructuring
                            raw_text = ""
                            if content_type in docx_mimes or ext == ".docx":
                                raw_text = extract_text_from_docx(payload)
                            elif content_type in xlsx_mimes or ext == ".xlsx":
                                raw_text = extract_text_from_xlsx(payload)
                            elif content_type in xls_mimes or ext == ".xls":
                                raw_text = extract_text_from_xls(payload)
                            else:
                                try:
                                    raw_text = payload.decode('utf-8', errors='ignore')
                                except Exception:
                                    try:
                                        raw_text = payload.decode('latin-1', errors='ignore')
                                    except Exception as de:
                                        print(f"[Attachment] Failed to decode text: {de}")

                            if raw_text.strip():
                                response = client.models.generate_content(
                                    model="gemini-2.5-flash",
                                    contents=(
                                        f"You are a purchasing document analyser. Extract ALL product names, item descriptions, "
                                        f"SKU codes, part numbers, quantities, and specifications mentioned in this text:\n\n"
                                        f"{raw_text}\n\n"
                                        f"Return them as a plain numbered list. Do NOT add any extra commentary or headings - "
                                        f"just list each line item or product request exactly as written. "
                                        f"If you cannot find any product items, reply with the single word: NONE."
                                    )
                                )
                                extracted = response.text.strip() if response.text else ""
                        break  # Success - exit retry loop

                    except Exception as api_err:
                        err_str = str(api_err)
                        is_transient = any(code in err_str for code in ["503", "429", "500", "UNAVAILABLE", "Resource has been exhausted"])
                        if is_transient and attempt < max_retries:
                            wait_sec = 2 ** attempt
                            print(f"[Attachment] Attempt {attempt}/{max_retries} failed (transient): {err_str[:80]}. Retrying in {wait_sec}s...")
                            time.sleep(wait_sec)
                        else:
                            raise
            except Exception as e:
                print(f"[Attachment] Gemini extraction failed for '{display_filename}': {e}")

        # 2. Local OCR Fallback for Images if Gemini is not available or failed
        if (not extracted or extracted.upper() == "NONE") and is_image:
            print(f"[Attachment] Gemini not available or failed. Falling back to local OCR for '{display_filename}'...")
            extracted = extract_text_via_local_ocr(payload, ext)

        # 3. Local Fallback for PDFs if Gemini is not available or failed
        is_pdf = ext == ".pdf" or content_type == "application/pdf"
        if (not extracted or extracted.upper() == "NONE") and is_pdf:
            print(f"[Attachment] Gemini not available or failed. Falling back to local PDF/OCR extraction for '{display_filename}'...")
            extracted = extract_text_from_pdf(payload)

        if extracted and extracted.upper() != "NONE":
            print(f"[Attachment] Extracted text from '{display_filename}':\n{extracted[:300]}...")
            extracted_texts.append(f"[From attachment '{display_filename}']:\n{extracted}")
        else:
            print(f"[Attachment] No product items found in '{display_filename}'.")

    return "\n\n".join(extracted_texts)

def get_graph_token_delegated(tenant_id, client_id, client_secret):
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    token_file = os.path.join(project_root, "data", f"outlook_tokens_{tenant_id}.json")
    if not os.path.exists(token_file):
        return None
        
    try:
        with open(token_file, "r", encoding="utf-8") as f:
            tokens = json.load(f)
    except Exception as e:
        print(f"[Outlook Auth] Error reading token file: {e}")
        return None
        
    mtime = os.path.getmtime(token_file)
    expires_in = tokens.get("expires_in", 3600)
    # 5-minute safety buffer
    if time.time() > mtime + expires_in - 300:
        print("[Outlook Auth] Access token expired or close to expiration. Refreshing...")
        refresh_token = tokens.get("refresh_token")
        if not refresh_token:
            print("[Outlook Auth] No refresh token found in token file.")
            return None
            
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        redirect_uri = "http://localhost:8080/api/outlook/callback"
        
        data = {
            "client_id": client_id,
            "client_secret": client_secret,
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "scope": "offline_access https://graph.microsoft.com/Mail.ReadWrite https://graph.microsoft.com/Mail.Send",
            "redirect_uri": redirect_uri
        }
        encoded_data = urllib.parse.urlencode(data).encode("utf-8")
        req = urllib.request.Request(token_url, data=encoded_data, headers={"Content-Type": "application/x-www-form-urlencoded"})
        
        try:
            with urllib.request.urlopen(req) as response:
                new_tokens = json.loads(response.read().decode("utf-8"))
                if "refresh_token" not in new_tokens:
                    new_tokens["refresh_token"] = refresh_token
                
                with open(token_file, "w", encoding="utf-8") as f:
                    json.dump(new_tokens, f, indent=2)
                return new_tokens.get("access_token")
        except Exception as re:
            print(f"[Outlook Auth] Error refreshing token: {re}")
            return None
    else:
        return tokens.get("access_token")

def get_graph_token(tenant_id, client_id, client_secret):
    # Try delegated token first
    del_token = get_graph_token_delegated(tenant_id, client_id, client_secret)
    if del_token:
        return del_token

    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default"
    }
    encoded_data = urllib.parse.urlencode(data).encode("utf-8")
    req = urllib.request.Request(url, data=encoded_data, headers={"Content-Type": "application/x-www-form-urlencoded"})
    try:
        with urllib.request.urlopen(req) as response:
            res_data = json.loads(response.read().decode("utf-8"))
            return res_data.get("access_token")
    except Exception as e:
        print(f"[Outlook Auth] Error acquiring token: {e}")
        return None

def fetch_outlook_messages(access_token, email_user):
    # Fetch recent messages from the last 60 minutes (regardless of read status)
    # and use processed_messages DB table as the deduplication gate.
    # This ensures emails read manually by the operator in Outlook are still captured.
    from datetime import datetime, timedelta, timezone
    since = (datetime.now(timezone.utc) - timedelta(minutes=60)).strftime("%Y-%m-%dT%H:%M:%SZ")
    filter_str = f"isDraft eq false and receivedDateTime ge {since}"
    encoded_filter = urllib.parse.quote(filter_str)
    url = f"https://graph.microsoft.com/v1.0/users/{email_user}/messages?$filter={encoded_filter}&$orderby=receivedDateTime desc&$top=50&$select=id,internetMessageId,subject,from,body,receivedDateTime,hasAttachments,isDraft"
    url = url.replace(" ", "%20")
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    })
    try:
        with urllib.request.urlopen(req) as response:
            res_data = json.loads(response.read().decode("utf-8"))
            return res_data.get("value", [])
    except Exception as e:
        # Fallback: if time-filter fails (e.g. older Graph API version), use isRead filter
        print(f"[Outlook Listener] Time-window filter failed ({e}), falling back to isRead filter.")
        url_fallback = f"https://graph.microsoft.com/v1.0/users/{email_user}/messages?$filter=isRead eq false and isDraft eq false&$orderby=receivedDateTime desc&$top=30"
        url_fallback = url_fallback.replace(" ", "%20")
        req2 = urllib.request.Request(url_fallback, headers={
            "Authorization": f"Bearer {access_token}",
            "Accept": "application/json"
        })
        with urllib.request.urlopen(req2) as resp2:
            res_data2 = json.loads(resp2.read().decode("utf-8"))
            return res_data2.get("value", [])

def fetch_outlook_attachments(access_token, email_user, message_id):
    url = f"https://graph.microsoft.com/v1.0/users/{email_user}/messages/{message_id}/attachments"
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json"
    })
    try:
        with urllib.request.urlopen(req) as response:
            res_data = json.loads(response.read().decode("utf-8"))
            return res_data.get("value", [])
    except Exception as e:
        print(f"[Outlook Mail] Error fetching attachments: {e}")
        return []

def _graph_request(url, access_token, method="GET", payload=None):
    """Minimal Microsoft Graph HTTP helper (urllib). Returns parsed JSON or {}."""
    data = json.dumps(payload).encode("utf-8") if payload is not None else None
    headers = {"Authorization": f"Bearer {access_token}"}
    if data is not None:
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    with urllib.request.urlopen(req) as resp:
        raw = resp.read().decode("utf-8")
        return json.loads(raw) if raw.strip() else {}


def _find_outlook_graph_id(access_token, email_user, internet_msg_id):
    """Resolve a Graph message resource id from an RFC Internet-Message-ID header."""
    if not internet_msg_id:
        return None
    val = internet_msg_id.strip()
    if not (val.startswith("<") and val.endswith(">")):
        val = f"<{val}>"
    filt = urllib.parse.quote(f"internetMessageId eq '{val}'", safe="")
    url = (f"https://graph.microsoft.com/v1.0/users/{email_user}/messages"
           f"?$filter={filt}&$select=id&$top=1")
    data = _graph_request(url, access_token, "GET")
    items = data.get("value", [])
    return items[0].get("id") if items else None


def _send_outlook_threaded_reply(access_token, email_user, internet_msg_id, subject, html_body, attachments):
    """Send `html_body` as a proper threaded reply to the customer's original email
    via Graph createReply (so it lands in the same conversation with correct
    In-Reply-To/References headers). Returns True on success, or False if the
    original message can't be located — in which case the caller falls back to a
    normal new-message send."""
    graph_id = _find_outlook_graph_id(access_token, email_user, internet_msg_id)
    if not graph_id:
        return False
    base = f"https://graph.microsoft.com/v1.0/users/{email_user}/messages"
    # 1. Draft a reply in the same conversation (inherits threading headers)
    draft = _graph_request(f"{base}/{graph_id}/createReply", access_token, "POST", {})
    draft_id = draft.get("id")
    if not draft_id:
        return False
    # 2. Get the draft's original content (to preserve email history/mail logs)
    try:
        draft_info = _graph_request(f"{base}/{draft_id}", access_token, "GET")
        original_history = draft_info.get("body", {}).get("content", "")
    except Exception as ge:
        print(f"[Outlook Send] Warning: Failed to retrieve draft body history: {ge}")
        original_history = ""

    # Combine our reply with the original thread history
    if original_history:
        # Prepend html_body to the original history, making sure we don't duplicate html/body tags
        clean_reply = html_body
        if "<body>" in clean_reply:
            parts = clean_reply.split("<body>")
            clean_reply = parts[1].split("</body>")[0] if len(parts) > 1 else clean_reply
        combined_body = f"<html><body>{clean_reply}<br><br>{original_history}</body></html>"
    else:
        combined_body = html_body

    # Overwrite the draft with combined body
    _graph_request(f"{base}/{draft_id}", access_token, "PATCH", {
        "body": {"contentType": "HTML", "content": combined_body},
    })
    # 3. Attach the PDF quote (and inline logo, if any)
    for att in attachments:
        _graph_request(f"{base}/{draft_id}/attachments", access_token, "POST", att)
    # 4. Send the threaded draft
    _graph_request(f"{base}/{draft_id}/send", access_token, "POST", {})
    return True


def send_outlook_mail(access_token, email_user, to_email, subject, html_body, pdf_path=None, logo_path=None, reply_to_internet_msg_id=None):
    url = f"https://graph.microsoft.com/v1.0/users/{email_user}/sendMail"
    
    to_recipients = [{"emailAddress": {"address": to_email}}]
    attachments = []
    
    # Attach Logo if exists and referenced
    if logo_path and os.path.exists(logo_path):
        try:
            with open(logo_path, "rb") as f:
                logo_bytes = base64.b64encode(f.read()).decode("utf-8")
            attachments.append({
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": os.path.basename(logo_path),
                "contentType": "image/png",
                "contentId": "company_logo",
                "isInline": True,
                "contentBytes": logo_bytes
            })
        except Exception as le:
            print(f"[Outlook Send] Warning: Failed to attach logo: {le}")
            
    # Attach PDF Quote if exists
    if pdf_path and os.path.exists(pdf_path):
        try:
            with open(pdf_path, "rb") as f:
                pdf_bytes = base64.b64encode(f.read()).decode("utf-8")
            attachments.append({
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": os.path.basename(pdf_path),
                "contentType": "application/pdf",
                "contentBytes": pdf_bytes
            })
        except Exception as pe:
            print(f"[Outlook Send] Warning: Failed to attach PDF: {pe}")

    # Preferred path: reply in the customer's original thread. Falls back to a
    # normal new-message send if the original message can't be found or the
    # reply flow errors, so a quote is never dropped.
    if reply_to_internet_msg_id:
        try:
            if _send_outlook_threaded_reply(access_token, email_user, reply_to_internet_msg_id, subject, html_body, attachments):
                print("[Outlook Send] Sent as threaded reply to the customer's original email.")
                return True
            print("[Outlook Send] Original message not found for threading; sending as a new message.")
        except Exception as te:
            print(f"[Outlook Send] Threaded reply failed ({te}); sending as a new message.")

    payload = {
        "message": {
            "subject": subject,
            "body": {
                "contentType": "HTML",
                "content": html_body
            },
            "toRecipients": to_recipients,
            "attachments": attachments
        }
    }
    
    req_body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=req_body, headers={
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    })
    
    try:
        with urllib.request.urlopen(req) as response:
            return True
    except Exception as e:
        print(f"[Outlook Send] Error sending email via Graph: {e}")
        if hasattr(e, "read"):
            try:
                print(e.read().decode("utf-8"))
            except Exception:
                pass
        return False

def mark_outlook_message_read(access_token, email_user, message_id):
    url = f"https://graph.microsoft.com/v1.0/users/{email_user}/messages/{message_id}"
    req_body = json.dumps({"isRead": True}).encode("utf-8")
    req = urllib.request.Request(url, data=req_body, headers={
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }, method="PATCH")
    try:
        with urllib.request.urlopen(req) as response:
            return True
    except Exception as e:
        print(f"[Outlook Mail] Error marking message {message_id} as read: {e}")
        return False

def extract_outlook_attachment_text(payload, filename, content_type):
    api_key = os.environ.get("GEMINI_API_KEY", "")
    has_gemini = api_key and not api_key.startswith("your_") and api_key.strip() != ""

    gemini_native_mimes = {
        "application/pdf": ".pdf",
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "image/tiff": ".tiff",
        "image/bmp": ".bmp",
    }

    docx_mimes = {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/docx": ".docx",
        "application/x-docx": ".docx"
    }

    xlsx_mimes = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/xlsx": ".xlsx",
        "application/x-xlsx": ".xlsx"
    }

    xls_mimes = {
        "application/vnd.ms-excel": ".xls",
        "application/xls": ".xls",
        "application/x-xls": ".xls"
    }

    ext = ""
    if filename:
        _, file_ext = os.path.splitext(filename)
        ext = file_ext.lower()
    else:
        if content_type in gemini_native_mimes:
            ext = gemini_native_mimes[content_type]
        elif content_type in docx_mimes:
            ext = ".docx"
        elif content_type in xlsx_mimes:
            ext = ".xlsx"
        elif content_type in xls_mimes:
            ext = ".xls"
        elif content_type.startswith("text/"):
            ext = ".txt"

    is_supported = (
        content_type in gemini_native_mimes or
        content_type in docx_mimes or
        content_type in xlsx_mimes or
        content_type in xls_mimes or
        (content_type and content_type.startswith("text/")) or
        (filename and filename.lower().endswith((".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif", ".tiff", ".bmp", ".docx", ".txt", ".rtf", ".html", ".xml", ".csv", ".xlsx", ".xls")))
    )

    if not is_supported or not payload:
        return ""

    display_filename = filename or f"attachment{ext}"
    print(f"[Attachment] Processing Outlook attachment: '{display_filename}' ({content_type}).")

    is_native_gemini = content_type in gemini_native_mimes or ext in [".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif", ".tiff", ".bmp"]
    is_image = ext in [".jpg", ".jpeg", ".png", ".webp", ".gif", ".tiff", ".bmp"]

    extracted = ""
    
    # 1. Try Gemini first if key is configured
    if has_gemini:
        try:
            from google import genai
            from google.genai import types as genai_types
            client = genai.Client(api_key=api_key)

            prompt = (
                "You are a purchasing document analyser. Extract ALL product names, item descriptions, "
                "SKU codes, part numbers, quantities, and specifications mentioned in this document. "
                "Return them as a plain numbered list. Do NOT add any extra commentary or headings - "
                "just list each line item or product request exactly as written. "
                "If you cannot find any product items, reply with the single word: NONE."
            )

            max_retries = 3
            for attempt in range(1, max_retries + 1):
                try:
                    if is_native_gemini:
                        response = client.models.generate_content(
                            model="gemini-2.5-flash",
                            contents=[
                                genai_types.Part.from_bytes(
                                    data=payload,
                                    mime_type=content_type if content_type in gemini_native_mimes else "application/pdf"
                                ),
                                prompt
                            ]
                        )
                        extracted = response.text.strip() if response.text else ""
                    else:
                        raw_text = ""
                        if content_type in docx_mimes or ext == ".docx":
                            raw_text = extract_text_from_docx(payload)
                        elif content_type in xlsx_mimes or ext == ".xlsx":
                            raw_text = extract_text_from_xlsx(payload)
                        elif content_type in xls_mimes or ext == ".xls":
                            raw_text = extract_text_from_xls(payload)
                        else:
                            try:
                                raw_text = payload.decode('utf-8', errors='ignore')
                            except Exception:
                                raw_text = payload.decode('latin-1', errors='ignore')

                        if not raw_text.strip():
                            return ""

                        response = client.models.generate_content(
                            model="gemini-2.5-flash",
                            contents=[
                                f"Document Content:\n{raw_text}\n\n{prompt}"
                            ]
                        )
                        extracted = response.text.strip() if response.text else ""
                    break
                except Exception as re:
                    print(f"[Attachment] Transient error on attempt {attempt}: {re}")
                    if attempt == max_retries:
                        raise re
                    time.sleep(1)
        except Exception as e:
            print(f"[Attachment] Error analyzing Outlook attachment {display_filename}: {e}")

    # 2. Local OCR Fallback for Images if Gemini is not available or failed
    if (not extracted or extracted.upper() == "NONE") and is_image:
        print(f"[Attachment] Gemini not available or failed. Falling back to local OCR for Outlook attachment '{display_filename}'...")
        extracted = extract_text_via_local_ocr(payload, ext)

    # 3. Local Fallback for PDFs if Gemini is not available or failed
    is_pdf = ext == ".pdf" or (content_type and content_type == "application/pdf")
    if (not extracted or extracted.upper() == "NONE") and is_pdf:
        print(f"[Attachment] Gemini not available or failed. Falling back to local PDF/OCR extraction for Outlook attachment '{display_filename}'...")
        extracted = extract_text_from_pdf(payload)

    if extracted and extracted != "NONE":
        return extracted

    return ""

def format_graph_datetime(dt_str):
    if not dt_str:
        return None
    try:
        dt_str = dt_str.replace("Z", "")
        if "." in dt_str:
            dt_str = dt_str.split(".")[0]
        from datetime import datetime, timedelta
        dt = datetime.strptime(dt_str, "%Y-%m-%dT%H:%M:%S")
        dt_ist = dt + timedelta(hours=5, minutes=30)
        return dt_ist.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return dt_str

def format_email_date(date_str):
    if not date_str:
        return None
    try:
        import email.utils
        from datetime import datetime, timedelta, timezone
        dt = email.utils.parsedate_to_datetime(date_str)
        dt_ist = dt.astimezone(timezone(timedelta(hours=5, minutes=30)))
        return dt_ist.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return date_str

def convert_html_to_plaintext(html_content):
    if not html_content:
        return ""
    import html as py_html
    # 1. Replace block line breaks
    text = re.sub(r'(?i)<br\s*/?>', '\n', html_content)
    # 2. Add tabs/spaces for table columns to prevent merging
    text = re.sub(r'(?i)</td\s*>', ' \t ', text)
    text = re.sub(r'(?i)</th\s*>', ' \t ', text)
    # 3. Add newlines for rows, paragraphs and block divisions
    text = re.sub(r'(?i)</tr\s*>', '\n', text)
    text = re.sub(r'(?i)</p\s*>', '\n\n', text)
    text = re.sub(r'(?i)</div\s*>', '\n', text)
    # 4. Strip all remaining HTML tags
    text = re.sub(r'<[^<]+?>', '', text)
    # 5. Unescape HTML entities
    text = py_html.unescape(text)
    # 6. Clean whitespace per line
    lines = [line.strip() for line in text.splitlines()]
    return '\n'.join(line for line in lines if line)

def poll_outlook_graph(catalog, crm_path, tenant_id, tenant_config, crm_emails, project_root):
    outlook_tenant_id = tenant_config.get("outlook_tenant_id")
    outlook_client_id = tenant_config.get("outlook_client_id")
    outlook_client_secret = tenant_config.get("outlook_client_secret")
    email_user = tenant_config.get("email_user")

    if not outlook_tenant_id or not outlook_client_id or not outlook_client_secret or not email_user:
        print(f"[Outlook Listener] Error: Missing Outlook configuration for tenant {tenant_id}")
        return

    token = get_graph_token(outlook_tenant_id, outlook_client_id, outlook_client_secret)
    if not token:
        try:
            from src.database_sqlite import update_service_status
            update_service_status("AUTH_FAILED", error_message="Failed to acquire Azure AD Graph token", tenant_id=tenant_id)
        except Exception:
            pass
        return

    try:
        from src.database_sqlite import update_service_status
        update_service_status("CONNECTED", tenant_id=tenant_id)
    except Exception:
        pass

    try:
        messages = fetch_outlook_messages(token, email_user)
    except Exception as e:
        print(f"[Outlook Mail] Error fetching messages: {e}")
        error_msg = str(e)
        if hasattr(e, "read"):
            try:
                error_msg += " - " + e.read().decode("utf-8")
            except Exception:
                pass
        
        status = "AUTH_FAILED" if "403" in error_msg or "401" in error_msg or "access" in error_msg.lower() else "ERROR"
        try:
            from src.database_sqlite import update_service_status
            update_service_status(status, error_message=f"Mailbox access error: {error_msg}", tenant_id=tenant_id)
        except Exception:
            pass
        return

    if not messages:
        try:
            update_service_status("IDLE", tenant_id=tenant_id)
        except Exception:
            pass
        return

    print(f"[Outlook Listener] Found {len(messages)} recent unread emails in Outlook inbox for {tenant_id}. Checking them.")

    for msg in messages:
        msg_id = msg.get("id")
        internet_msg_id = msg.get("internetMessageId")
        received_at_utc = msg.get("receivedDateTime")
        received_at = format_graph_datetime(received_at_utc)
        
        sender_email = msg.get("from", {}).get("emailAddress", {}).get("address", "")
        sender_name = msg.get("from", {}).get("emailAddress", {}).get("name", sender_email)
        
        if internet_msg_id:
            from src.database_sqlite import is_message_processed, log_processed_message
            if is_message_processed(internet_msg_id, tenant_id=tenant_id):
                mark_outlook_message_read(token, email_user, msg_id)
                continue
            
            # Atomic lock claim
            if not log_processed_message(internet_msg_id, "PROCESSING", received_at=received_at, tenant_id=tenant_id, customer_name=sender_name, customer_email=sender_email):
                print(f"[Outlook Listener] Race condition avoided: message {internet_msg_id} is already being processed by another instance.")
                mark_outlook_message_read(token, email_user, msg_id)
                continue
        subject = msg.get("subject")
        if not subject or not subject.strip() or subject.strip().lower() in ("(no subject)", "no subject", "(no-subject)", "no-subject"):
            subject = "Order Enquiry"

        if sender_email.lower() == email_user.lower():
            # ALWAYS skip emails sent from our own inbox — unconditionally.
            # Previously only checked subject keywords, allowing RE: threads to slip through and loop.
            print(f"[Outlook Listener] Ignored email from ourselves ({sender_email}) — self-sent loop prevention.")
            if internet_msg_id:
                from src.database_sqlite import log_processed_message
                log_processed_message(internet_msg_id, "SELF_SENT", received_at=received_at, tenant_id=tenant_id)
            mark_outlook_message_read(token, email_user, msg_id)
            continue

        body_content = msg.get("body", {}).get("content", "")
        is_html = msg.get("body", {}).get("contentType", "").lower() == "html"
        body = body_content
        if is_html:
            body = convert_html_to_plaintext(body_content)

        # Force has_attach to True to always fetch and scan for standard/inline attachments.
        # Graph API omits inline attachments (such as pasted screenshots) from 'hasAttachments'.
        has_attach = True
        attachment_text = ""
        clean_attach_text = ""
        
        if has_attach:
            attachments = fetch_outlook_attachments(token, email_user, msg_id)
            for att in attachments:
                name = att.get("name", "")
                content_bytes_b64 = att.get("contentBytes", "")
                content_type = att.get("contentType", "")
                if content_bytes_b64:
                    try:
                        file_data = base64.b64decode(content_bytes_b64)
                        text = extract_outlook_attachment_text(file_data, name, content_type)
                        if text:
                            attachment_text += f"\n[From attachment '{name}']:\n{text}\n"
                    except Exception as ae:
                        print(f"[Outlook Attachment] Warning: Failed to parse attachment {name}: {ae}")

            clean_attach_text = "" if not attachment_text else "\n".join(
                line for line in attachment_text.splitlines()
                if not line.startswith("[ATTACHMENT_FAILED:")
            ).strip()
            if clean_attach_text:
                print(f"[Attachment] Successfully extracted text from Outlook attachments. Merging with body.")
                body = (body + "\n\n" + clean_attach_text).strip()

        # ── Customer Reply Detection (Bug 1 + Bug 5 fix) ───────────────────────
        subj_lower = subject.lower().strip()
        # Widened is_reply_subject: any Re:/Fw: or contains QTN reference
        is_reply_subject = (
            subj_lower.startswith("re:") or
            subj_lower.startswith("fw:") or
            subj_lower.startswith("fwd:") or
            "reply" in subj_lower or
            "status" in subj_lower or
            "quote" in subj_lower or
            "follow" in subj_lower or
            bool(re.search(r'\[Quotation\s+#|QTN-[A-Z0-9\-]+', subj_lower, re.IGNORECASE))
        )
        
        # Step 1: Match QTN reference from subject line [Quotation #ID] or QTN-XXXXX
        qtn_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', subject, re.IGNORECASE)
        qtn_ref_extracted = None
        if qtn_match:
            qtn_ref_extracted = qtn_match.group(1).upper()
        else:
            qtn_match = re.search(r'(QTN-[A-Z0-9\-]+)', subject, re.IGNORECASE)
            if qtn_match:
                qtn_ref_extracted = qtn_match.group(1).upper()

        # Step 2: If no QTN in subject, scan the email body (thread history includes it)
        if not qtn_ref_extracted:
            body_qtn_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', body[:3000], re.IGNORECASE)
            if body_qtn_match:
                qtn_ref_extracted = body_qtn_match.group(1).upper()
                print(f"[Reply Detector] Found QTN reference in body: {qtn_ref_extracted}")
            else:
                body_qtn_match2 = re.search(r'(QTN-[A-Z0-9\-]+)', body[:3000], re.IGNORECASE)
                if body_qtn_match2:
                    qtn_ref_extracted = body_qtn_match2.group(1).upper()
                    print(f"[Reply Detector] Found QTN reference in body text: {qtn_ref_extracted}")

        # Step 3: If no QTN found anywhere, look up sender's most recent quotation (Bug 5: use LOWER for case-insensitive match)
        sender_qtn = None
        sender_email_normalized = sender_email.lower().strip() if sender_email else ""
        if not qtn_ref_extracted and sender_email_normalized:
            try:
                from src.database_sqlite import get_connection
                import datetime as _dt
                _rc = get_connection(tenant_id)
                # Look for any quotation from this sender in the last 30 days
                _30d_ago = (_dt.datetime.utcnow() - _dt.timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
                _row = _rc.execute(
                    "SELECT invoice_id FROM quotations WHERE LOWER(customer_email) = ? AND created_at >= ? ORDER BY rowid DESC LIMIT 1",
                    (sender_email_normalized, _30d_ago)
                ).fetchone()
                _rc.close()
                if _row:
                    sender_qtn = _row[0]
                    print(f"[Reply Detector] Sender {sender_email} has recent quotation {sender_qtn} (30-day fallback)")
            except Exception as _e:
                print(f"[Reply Detector] sender-quote lookup failed: {_e}")

        is_customer_reply = False
        qtn_ref = None
        candidate_qtn = qtn_ref_extracted or sender_qtn
        # If we have a reply-like subject OR any QTN reference found, check if quotation exists
        if candidate_qtn and (is_reply_subject or qtn_ref_extracted):
            try:
                from src.database_sqlite import get_connection
                _rc = get_connection(tenant_id)
                _exists = _rc.execute("SELECT 1 FROM quotations WHERE invoice_id = ?", (candidate_qtn,)).fetchone()
                _rc.close()
                if _exists:
                    is_customer_reply = True
                    qtn_ref = candidate_qtn
                    print(f"[Reply Detector] Confirmed customer reply for quotation {qtn_ref} from {sender_email}")
            except Exception as _e:
                print(f"[Reply Detector] quote existence check failed: {_e}")
        
        # Track whether we already logged this reply's customer message to avoid duplicates (Bug 3)
        _reply_customer_msg_logged = False

        # ── Fast Blocklist & Relevance checks ────────────────────────────────
        blocklist_result = fast_blocklist_check(sender_email, subject, crm_emails)
        if blocklist_result == "REJECT":
            print(f"[Outlook Filter] Skipped irrelevant email from {sender_email} (Subject: {subject})")
            if internet_msg_id:
                from src.database_sqlite import log_processed_message
                log_processed_message(internet_msg_id, "IRRELEVANT", received_at=received_at, tenant_id=tenant_id)
            mark_outlook_message_read(token, email_user, msg_id)
            continue

        # If it is a brand-new email thread (not a reply), even if they are in the CRM,
        # we must run the AI/relevance checks to filter out conversational messages.
        needs_relevance_check = (blocklist_result == "NEEDS_AI") or (blocklist_result == "ACCEPT_CRM" and not is_customer_reply)
        if needs_relevance_check:
            # Bug 11 fix: Never block emails with explicit QTN references in the subject
            subj_has_qtn_ref = bool(re.search(r'\[Quotation\s+#|QTN-[A-Z0-9\-]+', subject, re.IGNORECASE))
            if not subj_has_qtn_ref and not is_customer_reply and not is_subject_relevant(subject, sender_email, crm_emails, tenant_id=tenant_id):
                print(f"[Outlook Filter] Skipped irrelevant subject from {sender_email} (Subject: {subject})")
                if internet_msg_id:
                    from src.database_sqlite import log_processed_message
                    log_processed_message(internet_msg_id, "IRRELEVANT", received_at=received_at, tenant_id=tenant_id)
                mark_outlook_message_read(token, email_user, msg_id)
                continue
            
            ai_result = classify_and_extract(sender_email, subject, body)
            if ai_result and ai_result.get("intent") == "IRRELEVANT":
                # Even if AI says irrelevant, don't block if body contains a quotation reference
                body_has_qtn_ref = bool(re.search(r'\[Quotation\s+#|QTN-[A-Z0-9\-]+', body[:2000], re.IGNORECASE))
                if body_has_qtn_ref:
                    print(f"[AI Filter] AI said irrelevant but body has QTN reference — processing as thread reply. ({sender_email}, {subject})")
                else:
                    print(f"[AI Filter] Skipped irrelevant email from {sender_email} (Subject: {subject})")
                    if internet_msg_id:
                        from src.database_sqlite import log_processed_message
                        log_processed_message(internet_msg_id, "IRRELEVANT", received_at=received_at, tenant_id=tenant_id)
                    mark_outlook_message_read(token, email_user, msg_id)
                    continue

        # ── Activity Log: EMAIL_RECEIVED ─────────────────────────────────────
        try:
            from src.database_sqlite import log_activity
            _act_desc = f"Email received from {sender_name} <{sender_email}> (Subject: {subject[:80]})"
            log_activity(
                "EMAIL_RECEIVED",
                invoice_id=qtn_ref if is_customer_reply else None,
                customer_name=sender_name,
                customer_email=sender_email,
                description=_act_desc,
                tenant_id=tenant_id
            )
        except Exception:
            pass

        # ── Immediate DB Logging (Stage tracking) ──────────────────────────────
        if internet_msg_id:
            from src.database_sqlite import log_processed_message
            if is_customer_reply:
                log_processed_message(internet_msg_id, f"CUSTOMER_REPLIED:{qtn_ref}", received_at=received_at, customer_name=sender_name, customer_email=sender_email, tenant_id=tenant_id)
            else:
                log_processed_message(internet_msg_id, "NEW", received_at=received_at, customer_name=sender_name, customer_email=sender_email, tenant_id=tenant_id)

        # ── Thread context: load prior chat history for customer replies ───────
        prior_thread_context = None
        if is_customer_reply and qtn_ref and qtn_ref != "REPLIED":
            try:
                from src.database_sqlite import log_chat_msg, get_chat_history_for_context
                _reply = body or ""
                _cut = re.search(r'On\s+.+?\bwrote:', _reply, re.DOTALL)
                if _cut:
                    _reply = _reply[:_cut.start()]
                _reply_clean = _reply.strip() or "(customer replied)"
                # Bug 3 fix: Log customer reply here and mark as logged to prevent process_incoming_email from logging it again
                log_chat_msg(qtn_ref, "customer", _reply_clean, tenant_id=tenant_id)
                _reply_customer_msg_logged = True
                # Load full prior conversation for AI context
                prior_thread_context = get_chat_history_for_context(qtn_ref, tenant_id=tenant_id)
                if prior_thread_context:
                    print(f"[Thread Context] Loaded {len(prior_thread_context.splitlines())} prior messages for {qtn_ref}")
                # Activity log: CUSTOMER_REPLIED
                from src.database_sqlite import log_activity
                _reply_preview = _reply_clean[:120]
                log_activity(
                    "CUSTOMER_REPLIED",
                    invoice_id=qtn_ref,
                    customer_name=sender_name,
                    customer_email=sender_email,
                    description=f"Customer replied to {qtn_ref}: {_reply_preview}",
                    tenant_id=tenant_id
                )
            except Exception as _e:
                print(f"[Reply Detector] failed to log reply/load context: {_e}")
        elif not is_customer_reply:
            # New enquiry — no prior context needed
            prior_thread_context = None


        # ── Human Request Checking & Routing ───────────────────────────────
        is_human = is_human_request(body)
        if is_human:
            print(f"[Human Request] Customer requested human agent. Routing to Pending.")
            try:
                from src.database_sqlite import log_unmatched_item, log_activity
                u_id = log_unmatched_item(
                    customer_email=sender_email,
                    customer_name=sender_name,
                    original_body=f"HUMAN AGENT REQUESTED:\n{body}",
                    source="live_email",
                    tenant_id=tenant_id
                )
                proc_invoice_id = f"UNMATCHED_{u_id}"
                log_activity("HUMAN_AGENT_REQUESTED", invoice_id=proc_invoice_id,
                    customer_name=sender_name, customer_email=sender_email,
                    description="Customer requested a human sales agent", tenant_id=tenant_id)
            except Exception as ue:
                print(f"[Warning] Failed to log human request unmatched item: {ue}")
                proc_invoice_id = "UNMATCHED"
            
            reply_subject = clean_reply_subject(subject, is_unparsed=True)
            reply_body = (
                f"Dear {sender_name},\n\n"
                f"Thank you for contacting us. We have received your request and forwarded it "
                f"to a sales representative. A team member will contact you shortly to assist you.\n\n"
                f"Regards,\n"
                f"{tenant_config.get('business_name', 'Trofeo Hardware')} Sales Team"
            )
            plain_body = reply_body
            html_body = f"<html><body><p>{reply_body.replace(chr(10), '<br>')}</p></body></html>"
            pdf_path = None
            status = "UNPARSED_NOTICE"
        else:
            # Ingestion Mode is LIVE — pass thread context if this is a customer reply
            result = process_incoming_email(
                sender_email, subject, body, catalog, crm_path, "live", project_root,
                tenant_id=tenant_id,
                prior_thread_context=prior_thread_context,
                skip_initial_customer_log=_reply_customer_msg_logged
            )
            if len(result) == 5:
                reply_subject, reply_body_tuple, pdf_path, status, proc_invoice_id = result
            else:
                reply_subject, reply_body_tuple, pdf_path, status = result
                proc_invoice_id = getattr(result, "invoice_id", None)

            # Extract QTN invoice ID from the reply subject (e.g. "RE: ... [Quotation #QTN-00049]")
            if not proc_invoice_id:
                _qtn_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', reply_subject, re.IGNORECASE)
                if _qtn_match:
                    proc_invoice_id = _qtn_match.group(1)

            if status == "UNPARSED_NOTICE":
                print(f"[Unmatched] No valid SKU matches for live enquiry from {sender_email}. Logging to unmatched_items (No reply generated).")
                try:
                    from src.database_sqlite import log_unmatched_item, log_activity
                    # Prepend Subject, Sender, and Attachments to body for rich preview in simulator
                    full_body_log = f"Subject: {subject}\nSender: {sender_name} <{sender_email}>\n"
                    # Check if attachments exist in scope
                    if 'attachments' in locals() and attachments:
                        names = [att.get("name", "") for att in attachments if att.get("name")]
                        if names:
                            full_body_log += f"Attachments: {', '.join(names)}\n"
                    full_body_log += f"\n{body}"
                    
                    u_id = log_unmatched_item(
                        customer_email=sender_email,
                        customer_name=sender_name,
                        original_body=full_body_log.strip(),
                        source="live_email",
                        tenant_id=tenant_id
                    )
                    proc_invoice_id = f"UNMATCHED_{u_id}"
                    log_activity("UNMATCHED_ENQUIRY", invoice_id=proc_invoice_id,
                        customer_name=sender_name, customer_email=sender_email,
                        description=f"Items not found in catalog (Subject: {subject[:80]})",
                        tenant_id=tenant_id)
                except Exception as ue:
                    print(f"[Warning] Failed to log unmatched live enquiry: {ue}")
                    proc_invoice_id = "UNMATCHED"
                
                # Log processed message reference, mark as read, and skip sending reply
                if internet_msg_id:
                    from src.database_sqlite import log_processed_message
                    log_processed_message(internet_msg_id, proc_invoice_id, received_at=received_at, customer_name=sender_name, customer_email=sender_email, tenant_id=tenant_id)
                mark_outlook_message_read(token, email_user, msg_id)
                continue
            else:
                # Log QUOTE_GENERATED or other statuses to activity log
                try:
                    from src.database_sqlite import log_activity
                    _evt = "QUOTE_GENERATED" if status in ("QUOTE_GENERATED", "QUOTE_UPDATED") else status
                    log_activity(_evt, invoice_id=proc_invoice_id,
                        customer_name=sender_name, customer_email=sender_email,
                        description=f"Status: {status} for {proc_invoice_id} (Subject: {subject[:70]})",
                        tenant_id=tenant_id)
                except Exception:
                    pass

            if isinstance(reply_body_tuple, tuple):
                plain_body, html_body = reply_body_tuple
            else:
                plain_body = reply_body_tuple
                html_body = f"<html><body><p>{plain_body.replace(chr(10), '<br>')}</p></body></html>"

        # Check reply mode setting
        from src.database_sqlite import get_setting, update_quotation_status, log_chat_msg
        reply_mode = get_setting("reply_mode", "auto", tenant_id)
        
        # In manual mode, we hold any generated reply or quote (including thread replies)
        is_held_draft = (status in ("QUOTE_GENERATED", "QUOTE_UPDATED", "CONVERSATIONAL_REPLY", "CUSTOMER_REPLIED") or is_customer_reply)

        if reply_mode == "manual" and is_held_draft:
            # Bug 7 fix: For thread replies, always use qtn_ref as the fallback invoice ID
            ref_id = proc_invoice_id or qtn_ref
            if ref_id:
                update_quotation_status(ref_id, "PENDING_REVIEW", tenant_id=tenant_id)
            
            # Log the draft reply so we can show it to the operator
            draft_msg = f"Subject: {reply_subject}\n\n{plain_body}"
            log_chat_msg(ref_id or "MANUAL", "DRAFT_BOT", draft_msg, tenant_id=tenant_id)
            
            print(f"[Outlook Mailer] Manual mode active. Holding reply draft {ref_id} for approval.")
            sent = True
        else:
            sent = send_outlook_mail(
                token, email_user, sender_email, reply_subject, html_body,
                pdf_path=pdf_path, logo_path=tenant_config.get("company_logo_path"),
                reply_to_internet_msg_id=internet_msg_id
            )

        if sent:
            if reply_mode == "manual" and is_held_draft:
                print(f"[Outlook Mailer] Held draft response for manual review: {proc_invoice_id or qtn_ref}")
            else:
                print(f"[Outlook Mailer] Successfully sent reply to {sender_email} (Subject: {reply_subject})")
            # Bug 12 fix: Always log EMAIL_SENT to activity log regardless of status (including UNPARSED_NOTICE)
            try:
                from src.database_sqlite import log_activity
                _sent_inv = proc_invoice_id or (f"CUSTOMER_REPLIED:{qtn_ref}" if is_customer_reply and qtn_ref else None)
                log_activity("EMAIL_SENT", invoice_id=_sent_inv,
                    customer_name=sender_name, customer_email=sender_email,
                    description=f"Reply sent to {sender_email} — Subject: {reply_subject[:80]}",
                    tenant_id=tenant_id)
            except Exception:
                pass
            if is_customer_reply and qtn_ref and qtn_ref != "REPLIED":
                try:
                    from src.database_sqlite import log_chat_msg
                    # Log the bot's reply to the thread timeline
                    bot_msg = (plain_body or "").strip() or f"AI responded to customer's reply for {qtn_ref}."
                    log_chat_msg(qtn_ref, "BOT", bot_msg, tenant_id=tenant_id)
                except Exception as _ce:
                    print(f"[Reply Detector] failed to log bot reply to timeline: {_ce}")
            if internet_msg_id:
                from src.database_sqlite import log_processed_message
                if is_customer_reply:
                    log_invoice_ref = f"CUSTOMER_REPLIED:{qtn_ref}"
                elif status == "UNPARSED_NOTICE":
                    log_invoice_ref = proc_invoice_id or "UNMATCHED"
                else:
                    log_invoice_ref = proc_invoice_id or status
                log_processed_message(internet_msg_id, log_invoice_ref, received_at=received_at, customer_name=sender_name, customer_email=sender_email, tenant_id=tenant_id)
        else:
            print(f"[Outlook Mailer] Error sending reply to {sender_email}")

        mark_outlook_message_read(token, email_user, msg_id)

    try:
        update_service_status("IDLE", tenant_id=tenant_id)
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# AI-Powered Email Classification (Tier 1 + Tier 2)
# ─────────────────────────────────────────────────────────────────────────────

_gemini_client_cache = {}

def _get_gemini_client():
    """Returns a cached Gemini client instance, or None if API key is unavailable."""
    if "client" in _gemini_client_cache and _gemini_client_cache["client"] is not None:
        return _gemini_client_cache["client"]
    
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key or api_key.strip() == "" or api_key.startswith("your_"):
        _gemini_client_cache["client"] = None
        return None
    
    try:
        from google import genai
        client = genai.Client(api_key=api_key)
        _gemini_client_cache["client"] = client
        return client
    except Exception as e:
        print(f"[AI Classifier] Failed to initialize Gemini client: {e}")
        _gemini_client_cache["client"] = None
        return None


def fast_blocklist_check(sender, subject, crm_emails):
    """
    Tier 1 fast pre-filter. Checks blocklists and CRM match without any API calls.
    Returns: "REJECT", "ACCEPT_CRM", "ACCEPT_THREAD", or "NEEDS_AI"
    """
    sender_lower = sender.lower().strip()
    subject_lower = subject.lower().strip()
    
    # Extract email address from sender
    from email.utils import parseaddr
    _, email_addr = parseaddr(sender)
    email_addr_lower = (email_addr or sender).lower().strip()
    
    # 1. Blocklist check for automated/promotional senders (unless registered in CRM)
    if email_addr_lower not in crm_emails and sender_lower not in crm_emails:
        system_sender_keywords = [
            "mailer-daemon", "daemon", "postmaster", "bounce", "noreply", "no-reply", 
            "donotreply", "do-not-reply", "newsletter", "notification", "alert", 
            "support", "marketing", "promo", "digest", "feedback", "updates", "news", 
            "community", "info@", "hello@", "welcome@", "billing@", "invoice@", "receipt@",
            "delivery@", "shipping@", "track@", "status@"
        ]
        if any(kw in sender_lower for kw in system_sender_keywords):
            print(f"[Blocklist] REJECT: Sender {sender} matched automated/system email blocklist.")
            return "REJECT"
            
        system_display_names = [
            "subsystem", "daemon", "service", "system", "mindvalley", "apollo", 
            "github", "gitlab", "google", "microsoft", "zoom", "slack", 
            "linkedin", "facebook", "twitter", "instagram", "amazon", "paypal", 
            "stripe"
        ]
        display_name = ""
        if "<" in sender:
            display_name = sender.split("<")[0].lower().strip()
        else:
            display_name = sender_lower
        display_name = display_name.replace('"', '').replace("'", "").strip()
        if any(kw in display_name for kw in system_display_names):
            print(f"[Blocklist] REJECT: Display name '{display_name}' matched automated/system blocklist.")
            return "REJECT"

        system_subject_keywords = [
            "delivery status", "undeliverable", "returned mail", "bounce", "failure notice", 
            "out of office", "auto-reply", "auto reply", "vacation", "spam", "unsubscribed", 
            "newsletter", "digest", "invoice paid", "payment receipt", "receipt for", 
            "welcome to", "verification code", "otp", "security alert", "password reset"
        ]
        if any(kw in subject_lower for kw in system_subject_keywords):
            print(f"[Blocklist] REJECT: Subject '{subject}' matched automated/bounce subject blocklist.")
            return "REJECT"

    # 2. Check if sender is a registered CRM client
    if email_addr_lower in crm_emails or sender_lower in crm_emails:
        print(f"[Blocklist] ACCEPT_CRM: Sender {sender} is a registered CRM client.")
        return "ACCEPT_CRM"
        
    # 3. Check if subject has Quotation ID reference
    if "quotation #" in subject_lower or "quote #" in subject_lower:
        print(f"[Blocklist] ACCEPT_THREAD: Subject refers to an active quotation reference.")
        return "ACCEPT_THREAD"
    
    return "NEEDS_AI"


def is_subject_relevant(subject, sender, crm_emails, tenant_id=None):
    """
    Checks if an incoming email subject or sender suggests it is a quote request
    or thread reply, avoiding calling Gemini on completely irrelevant emails.
    """
    subject_lower = subject.lower().strip()
    sender_lower = sender.lower().strip()
    
    # 1. Thread replies are always relevant
    if "quotation #" in subject_lower or "quote #" in subject_lower or "[quotation" in subject_lower:
        return True
        
    # 2. Registered CRM clients are always relevant
    if sender_lower in crm_emails:
        return True
        
    # 3. Subject keyword matching (whole word boundary checks) loaded from database
    try:
        from src.database_sqlite import get_training_keywords
        keywords = get_training_keywords(tenant_id)
    except Exception as e:
        print(f"[Warning] Failed to load training keywords from DB: {e}")
        keywords = [
            "quote", "quotation", "enquiry", "enquiries", "inquiry", "inquiries", 
            "pricing", "need", "needed", "material", "materials", "mtl", "mtls", 
            "rfq", "purchase", "order", "price", "prices", "request", "hardware", 
            "fastener", "fasteners", "match", "estimate", "estimating", "invoice",
            "requisition", "req", "items", "slip", "rfp", "vendor", "signoff",
            "welcome", "discussion", "onboarding", "agreement", "contract", "sign",
            "setup", "register", "registration", "details", "proposal", "proposals", 
            "commercial", "commercials", "offer", "offers", "rate", "rates", "bid", 
            "bids", "tender", "tenders"
        ]
        
    if any(re.search(r'\b' + re.escape(kw) + r'\b', subject_lower) for kw in keywords):
        return True
        
    # Substring checks for compound tokens
    if any(kw in subject_lower for kw in keywords if len(kw) >= 3):
        return True
        
    return False


def call_with_retry(api_func, max_retries=3, initial_delay=1.0, backoff_factor=2.0):
    """
    Executes an API function with exponential backoff and jitter for rate-limits or transient errors.
    """
    import random
    import time
    delay = initial_delay
    for attempt in range(max_retries + 1):
        try:
            return api_func()
        except Exception as e:
            err_msg = str(e).lower()
            # Catch HTTP 429 rate limit or 503/500 overload errors
            is_rate_limit = "429" in err_msg or "rate" in err_msg or "resource exhausted" in err_msg
            is_transient = "503" in err_msg or "overloaded" in err_msg or "unavailable" in err_msg or "connection" in err_msg
            
            if (is_rate_limit or is_transient) and attempt < max_retries:
                jitter = random.uniform(0.8, 1.2)
                sleep_time = delay * jitter
                print(f"[AI Classifier] API failed: {e}. Retrying in {sleep_time:.2f}s (attempt {attempt + 1}/{max_retries})...")
                time.sleep(sleep_time)
                delay *= backoff_factor
            else:
                raise e

def classify_and_extract(sender, subject, body):
    """
    Tier 2 AI-powered filter. Uses Gemini 2.5 Flash to classify email intent
    and extract product items in a single API call.
    
    Returns: dict with keys 'intent', 'items', 'confidence'
             or None if API is unavailable (signals caller to fall back to rule-based).
    """
    client = _get_gemini_client()
    if not client:
        return None  # Fallback signal
    
    try:
        prompt = f"""You are an email classifier for a hardware and industrial supplies company.
Analyze the following email and determine if it is a genuine product enquiry or purchase request.

Email From: {sender}
Email Subject: {subject}
Email Body:
---
{body[:3000]}
---

Respond with ONLY a valid JSON object (no markdown, no code fences, no explanation):
{{"intent": "<PRODUCT_ENQUIRY|NEGOTIATION|CONVERSATION|IRRELEVANT>", "items": [{{"product": "<product name as written by customer>", "quantity": <number>}}], "confidence": <0.0 to 1.0>}}

Classification rules:
- "PRODUCT_ENQUIRY": Customer is requesting pricing, quotation, or availability of specific products
- "NEGOTIATION": Customer is negotiating price or discount on an existing quotation
- "CONVERSATION": Customer is replying to an existing conversation or confirming an order
- "IRRELEVANT": Newsletter, spam, out-of-office, promotional, job application, personal chat, or any non-purchase-related email
- For IRRELEVANT emails, set items to an empty list []
- For PRODUCT_ENQUIRY, extract every product mentioned with its quantity (default to 1 if not specified)
- Do NOT invent or hallucinate products not mentioned in the email body"""
        
        def _execute_api():
            return client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt
            )
            
        response = call_with_retry(_execute_api)
        
        result_text = response.text.strip()
        # Clean markdown fences if Gemini wraps the response
        if result_text.startswith("```"):
            result_text = re.sub(r'^```(?:json)?\s*', '', result_text)
            result_text = re.sub(r'\s*```$', '', result_text)
        
        result = json.loads(result_text)
        
        # Validate required structure
        if "intent" not in result:
            return None
        if "items" not in result:
            result["items"] = []
        if "confidence" not in result:
            result["confidence"] = 0.5
            
        print(f"[AI Classifier] Intent: {result['intent']}, Items: {len(result['items'])}, Confidence: {result['confidence']:.2f}")
        return result
        
    except Exception as e:
        print(f"[AI Classifier] Gemini classification failed: {e}. Falling back to rule-based filter.")
        return None


def is_email_relevant(sender, subject, body, catalog, crm_emails, attachment_text="", email_has_attachments=False):
    """
    Checks if an incoming email is relevant to Trofeo Hardware.
    An email is relevant if:
    1. It passes automated/system/marketing blocklists.
    2. Sender's email is in our CRM database.
    3. Subject contains a quotation reference (e.g. [Quotation #1635]) or is a reply.
    4. Body contains any SKU ID from our catalog.
    5. Subject matches enquiry interest keywords AND body contains catalog product keywords.
    6. Body contains catalog product keywords AND quantity patterns.
    7. Attachment text (extracted via Gemini) contains product keywords.
    8. Email has attachments AND subject suggests product enquiry.
    """
    sender_lower = sender.lower().strip()
    subject_lower = subject.lower().strip()
    combined_body = ((body or "") + "\n" + (attachment_text or "")).strip()
    body_lower = combined_body.lower()
    
    # 1. Blocklist check for automated/promotional senders & bounce daemons (unless registered in CRM)
    if sender_lower not in crm_emails:
        system_sender_keywords = [
            "mailer-daemon", "daemon", "postmaster", "bounce", "noreply", "no-reply", 
            "donotreply", "do-not-reply", "newsletter", "notification", "alert", 
            "support", "marketing", "promo", "digest", "feedback", "updates", "news", 
            "community", "info@", "hello@", "welcome@", "billing@", "invoice@", "receipt@",
            "delivery@", "shipping@", "track@", "status@"
        ]
        if any(kw in sender_lower for kw in system_sender_keywords):
            print(f"[Email Filter] REJECT: Sender {sender} matched automated/system email blocklist.")
            return False
            
        system_display_names = [
            "subsystem", "daemon", "service", "system", "mindvalley", "apollo", 
            "odoo", "github", "gitlab", "google", "microsoft", "zoom", "slack", 
            "linkedin", "facebook", "twitter", "instagram", "amazon", "paypal", 
            "stripe"
        ]
        display_name = ""
        if "<" in sender:
            display_name = sender.split("<")[0].lower().strip()
        else:
            display_name = sender_lower
        display_name = display_name.replace('"', '').replace("'", "").strip()
        if any(kw in display_name for kw in system_display_names):
            print(f"[Email Filter] REJECT: Display name '{display_name}' matched automated/system blocklist.")
            return False

        system_subject_keywords = [
            "delivery status", "undeliverable", "returned mail", "bounce", "failure notice", 
            "out of office", "auto-reply", "auto reply", "vacation", "spam", "unsubscribed", 
            "newsletter", "digest", "invoice paid", "payment receipt", "receipt for", 
            "welcome to", "verification code", "otp", "security alert", "password reset"
        ]
        if any(kw in subject_lower for kw in system_subject_keywords):
            print(f"[Email Filter] REJECT: Subject '{subject}' matched automated/bounce subject blocklist.")
            return False

    # 2. Check if sender is a registered CRM client
    if sender_lower in crm_emails:
        print(f"[Email Filter] MATCH: Sender {sender} is a registered CRM client.")
        return True
        
    # 3. Check if subject has Quotation ID reference (e.g. Quotation #1635)
    if "quotation #" in subject_lower or "quote #" in subject_lower:
        print(f"[Email Filter] MATCH: Subject refers to an active quotation reference.")
        return True
        
    # 4. Check if body contains any SKU ID from the catalog
    for sku in catalog.skus:
        sku_id = sku.get("sku_id", "")
        if sku_id and sku_id.lower() in body_lower:
            print(f"[Email Filter] MATCH: Body contains SKU ID '{sku_id}'.")
            return True
            
    # 5. Check if body contains catalog product keywords
    product_keywords = [
        # Plumbing / Fittings
        "elbow", "fitting", "coupling", "joint", "valve", "gate valve", "pipe", "nipple", "adapter", 
        "teflon", "ptfe", "gasket", "seal", "hose", "pvc", "brass", "copper", "union", "plumbing",
        # Fasteners
        "bolt", "nut", "screw", "washer", "fastener", "rivet", "anchor", "nail", "thread", "threaded",
        # Tools
        "hammer", "claw hammer", "level", "spirit level", "tape", "seal tape", "brush", "paint brush", 
        "tool", "tools", "wrench", "screwdriver", "pliers", "saw", "drill", "lubricant", "wd-40", "spray",
        # General Hardware Materials
        "hardware", "steel", "metal", "iron", "zinc", "silicone", "sealant", "adhesive", "mounting"
    ]
    body_has_product_keywords = False
    matched_pk = ""
    for pk in product_keywords:
        if re.search(r'\b' + re.escape(pk) + r'\b', body_lower) or re.search(r'\b' + re.escape(pk) + r'\b', subject_lower):
            body_has_product_keywords = True
            matched_pk = pk
            break
            
    # 6. Check general sales/RFQ keywords in subject
    interest_keywords = [
        "enquiry", "inquiry", "order", "rfq", "quote", "purchase", "materials", "material", "request", 
        "quotation", "price", "pricing", "cost", "negotiate", "discount", "estimate", "invoice", "proforma", 
        "po", "billing", "rate", "rates", "rfp", "specsheet", "specification", "specifications", "tender", 
        "supplier", "vendor", "delivery", "leadtime", "lead time", "payment", "terms", "requisition", "valves",
        "fitting", "fittings", "bolts", "nuts", "screws", "fasteners", "washers"
    ]
    subject_matches_interest = any(kw in subject_lower for kw in interest_keywords)
    
    if subject_matches_interest and body_has_product_keywords:
        print(f"[Email Filter] MATCH: Subject matched interest keywords and body/subject matched product keyword '{matched_pk}'.")
        return True
        
    # 7. Check if body matches catalog product keywords and quantity patterns
    has_quantity_pattern = False
    if re.search(r'\b\d+(?:\.\d+)?\s*(?:x|units?|rolls?|pcs?|pieces?|cans?|lengths?|boxes?|bottles?|counts?|nos?|numbers?|qty|packet|pack|pkt|bags?|mts?|meters?|mtrs?)\b', combined_body, re.IGNORECASE):
        has_quantity_pattern = True
        
    if body_has_product_keywords and has_quantity_pattern:
        print(f"[Email Filter] MATCH: Body matched product keywords and quantity pattern.")
        return True

    # 7.5. Subject matches interest and body has quantity patterns (custom RFQ for unrecognized items)
    if subject_matches_interest and has_quantity_pattern:
        print(f"[Email Filter] MATCH: Subject matches interest keywords and body contains quantity patterns.")
        return True

    # 8. If attachment text was extracted and contains product-relevant info, pass it through
    if attachment_text and attachment_text.strip() and attachment_text.upper() != "NONE":
        print(f"[Email Filter] MATCH: Attachment text extracted — treating as potential product enquiry.")
        return True

    # 9. Email has attachments AND subject suggests it is a product enquiry with document
    #    (e.g. customer sends image/PDF of product without textual body)
    if email_has_attachments:
        attachment_subject_hints = [
            "product", "item", "want", "need", "require", "attached", "attach", 
            "image", "photo", "picture", "document", "doc", "file", "list",
            "buy", "purchase", "order", "get", "stock", "available", "price"
        ]
        if any(hint in subject_lower for hint in attachment_subject_hints):
            print(f"[Email Filter] MATCH: Email has attachment + subject hints at product enquiry ('{subject}').")
            return True

    return False

def send_master_notification(smtp_server, smtp_port, email_user, email_pass, master_email, subject, body_text):
    """Sends a notification email to the Master/Admin user using the SMTP settings."""
    if not master_email or not email_user or not email_pass:
        return
        
    try:
        msg = MIMEMultipart()
        msg["From"] = email_user
        msg["To"] = master_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body_text, 'plain'))
        
        # Support multiple comma-separated master emails
        recipients = [r.strip() for r in master_email.split(",") if r.strip()]
        
        if smtp_port == 465:
            server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        else:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
        server.login(email_user, email_pass)
        server.sendmail(email_user, recipients, msg.as_string())
        server.close()
        print(f"[Master Notification] Sent notification to {recipients} (Subject: {subject})")
    except Exception as e:
        print(f"[Warning] Failed to send master notification: {e}")

def send_unmatched_products_alert(smtp_server, smtp_port, email_user, email_pass, master_email, customer_name, customer_email, customer_phone, original_subject, unmatched_lines):
    """Sends an email notification to the Master/Admin when the customer requests a product not in the catalog."""
    if not master_email or not email_user or not email_pass:
        return
        
    try:
        subject = f"[ATTENTION] Unmatched Products in Enquiry from {customer_name}"
        
        items_list_str = ""
        for idx, line in enumerate(unmatched_lines, 1):
            items_list_str += f"{idx}. Line: \"{line['original_line']}\"\n   Detected Query: \"{line['parsed_query']}\" | Qty: {line['quantity']}\n"
            
        body_text = (
            f"Dear Master User,\n\n"
            f"The auto-bot processed an enquiry from the customer. While a partial quote may have been "
            f"generated/updated for matched items, the customer is asking for the following products "
            f"that are NOT in our master catalog:\n\n"
            f"{items_list_str}\n"
            f"Customer Details:\n"
            f"- Name: {customer_name}\n"
            f"- Email: {customer_email}\n"
            f"- Contact Number: {customer_phone}\n\n"
            f"Original Enquiry Subject: {original_subject}\n\n"
            f"Please check if you need to add these products to the catalog or contact the customer to arrange a manual quote.\n\n"
            f"Regards,\n"
            f"Trofeo Auto-bot"
        )
        
        send_master_notification(smtp_server, smtp_port, email_user, email_pass, master_email, subject, body_text)
    except Exception as e:
        print(f"[Warning] Failed to send unmatched products alert: {e}")

def adjust_quantities_by_stock(matched_lines, catalog, cap_by_stock=True, invoice_id=None, customer_name=None, customer_email=None, customer_phone=None, tenant_id=None):
    deficit_lines = []
    for line in matched_lines:
        sku_id = line.get("matched_sku_id")
        if sku_id and sku_id != "UNKNOWN":
            sku_item = next((s for s in catalog.skus if s["sku_id"] == sku_id), None)
            if sku_item:
                try:
                    stock_avail = int(sku_item.get("stock", 0))
                except Exception:
                    stock_avail = 100
                
                # Retrieve or initialize the original requested quantity
                if "original_requested_qty" not in line:
                    line["original_requested_qty"] = line["quantity"]
                
                req_qty = line["original_requested_qty"]
                line["stock_avail"] = stock_avail
                
                if req_qty > stock_avail:
                    if cap_by_stock:
                        line["quantity"] = 0 # Exclude from quotation entirely since requested qty exceeds stock
                    else:
                        line["quantity"] = req_qty
                    line["deficit"] = req_qty - stock_avail
                    deficit_lines.append(line)
                else:
                    line["quantity"] = req_qty
                    line["deficit"] = 0
            else:
                line["stock_avail"] = 100
                line["deficit"] = 0
        else:
            line["deficit"] = 0

    if deficit_lines and invoice_id:
        from src.database_sqlite import log_deficit
        for line in deficit_lines:
            try:
                log_deficit(
                    invoice_id=invoice_id,
                    sku_id=line["matched_sku_id"],
                    sku_name=line.get("matched_sku_name", "UNKNOWN"),
                    requested_qty=line["original_requested_qty"],
                    available_qty=line.get("stock_avail", 0),
                    deficit_qty=line["deficit"],
                    customer_name=customer_name or "Walk-in Retail Client",
                    customer_email=customer_email or "",
                    customer_phone=customer_phone or "—",
                    tenant_id=tenant_id
                )
            except Exception as e:
                print(f"[Warning] Failed to log deficit inside adjust_quantities_by_stock: {e}")

    return deficit_lines


def send_deficit_purchase_order_alert(smtp_server, smtp_port, email_user, email_pass, master_email, customer_name, customer_email, customer_phone, original_subject, deficit_lines):
    """Sends an email notification to the Master when customer requests quantities exceeding available stock."""
    if not master_email or not email_user or not email_pass:
        return
        
    try:
        subject = f"[PURCHASE ORDER REQUIRED] Stock Deficit for Customer {customer_name}"
        
        items_list_str = ""
        for idx, line in enumerate(deficit_lines, 1):
            items_list_str += (
                f"{idx}. SKU: {line['matched_sku_id']} ({line['matched_sku_name']})\n"
                f"   Requested Qty: {line['original_requested_qty']} | Available Qty: {line.get('stock_avail', 0)} | Deficit Qty: {line['deficit']}\n"
            )
            
        body_text = (
            f"Dear Master User,\n\n"
            f"An enquiry was processed for the customer, but some items had insufficient stock.\n"
            f"The customer has been notified of the available quantities, and the unavailable items "
            f"were excluded from their quotation.\n\n"
            f"Please generate a Purchase Order (PO) to fulfill the following deficit quantities:\n\n"
            f"{items_list_str}\n"
            f"Customer Details:\n"
            f"- Name: {customer_name}\n"
            f"- Email: {customer_email}\n"
            f"- Contact Number: {customer_phone}\n\n"
            f"Original Enquiry Subject: {original_subject}\n\n"
            f"Regards,\n"
            f"Trofeo Auto-bot"
        )
        
        send_master_notification(smtp_server, smtp_port, email_user, email_pass, master_email, subject, body_text)
    except Exception as e:
        print(f"[Warning] Failed to send deficit PO alert: {e}")

def send_quotation_email_to_customer(tenant_id, customer_email, reply_subject, plain_body, html_body, pdf_path, reply_to_msg_id=None):
    """
    Sends the quotation email containing the PDF attachment to the customer.
    If the tenant Outlook settings are set, sends via MS Graph.
    If the tenant SMTP settings are set, sends via SMTP.
    Otherwise, it writes to mock_outbox.
    """
    from src.tenants import get_tenant_config
    tenant_config = get_tenant_config(tenant_id)
    email_user = tenant_config.get("email_user")
    email_pass = tenant_config.get("email_pass")
    smtp_server = tenant_config.get("smtp_server", "smtp.gmail.com")
    
    outlook_tenant_id = tenant_config.get("outlook_tenant_id")
    outlook_client_id = tenant_config.get("outlook_client_id")
    outlook_client_secret = tenant_config.get("outlook_client_secret")
    
    # ── Auto-resolve reply_to_msg_id if not explicitly passed ──────────────
    if not reply_to_msg_id:
        extracted_inv_id = None
        if pdf_path:
            basename = os.path.basename(pdf_path)
            inv_match = re.search(r'Quote_([A-Za-z0-9\-]+)\.pdf', basename)
            if inv_match:
                extracted_inv_id = inv_match.group(1)
        if not extracted_inv_id:
            inv_match = re.search(r'\[Quotation\s+#([A-Za-z0-9\-]+)\]', reply_subject, re.IGNORECASE)
            if inv_match:
                extracted_inv_id = inv_match.group(1)
        
        if extracted_inv_id:
            try:
                from src.database_sqlite import get_latest_message_id
                reply_to_msg_id = get_latest_message_id(extracted_inv_id, tenant_id=tenant_id)
            except Exception as e:
                print(f"[Warning] Failed to look up Message-ID for threading: {e}")
    # ──────────────────────────────────────────────────────────────────────────

    if email_user and outlook_client_secret:
        try:
            token = get_graph_token(outlook_tenant_id, outlook_client_id, outlook_client_secret)
            if token:
                project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                logo_file_path = tenant_config.get("company_logo_path")
                if logo_file_path:
                    if not os.path.isabs(logo_file_path):
                        logo_file_path = os.path.join(project_root, logo_file_path)
                else:
                    logo_file_path = find_company_logo(project_root)
                
                sent = send_outlook_mail(
                    token,
                    email_user,
                    customer_email,
                    reply_subject,
                    html_body,
                    pdf_path=pdf_path,
                    logo_path=logo_file_path,
                    reply_to_internet_msg_id=reply_to_msg_id
                )
                if sent:
                    print(f"[Outlook Mailer] Successfully sent email to {customer_email} (Subject: {reply_subject})")
                    return True
                else:
                    print(f"[Outlook Mailer] Failed to send email to {customer_email} via Graph.")
            else:
                print(f"[Outlook Mailer] Error: Failed to acquire Graph token for sending.")
        except Exception as ge:
            print(f"[Outlook Mailer] Error sending email via Graph: {ge}")
            
    is_mock = not email_user or not email_pass or email_user.strip() == "" or email_user.startswith("your_")
    
    if is_mock:
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if tenant_id and tenant_id != "default":
            outbox_dir = os.path.join(project_root, "mock_outbox", tenant_id)
        else:
            outbox_dir = os.path.join(project_root, "mock_outbox")
        os.makedirs(outbox_dir, exist_ok=True)
        
        invoice_id = "UNKNOWN"
        if pdf_path:
            basename = os.path.basename(pdf_path)
            inv_match = re.search(r'Quote_([A-Za-z0-9\-]+)\.pdf', basename)
            if inv_match:
                invoice_id = inv_match.group(1)
        if invoice_id == "UNKNOWN":
            inv_match = re.search(r'\[Quotation\s+#([A-Za-z0-9\-]+)\]', reply_subject, re.IGNORECASE)
            if inv_match:
                invoice_id = inv_match.group(1)
                
        reply_filename = f"Quote_{invoice_id}_reply.txt" if invoice_id != "UNKNOWN" else f"mock_reply_{int(time.time())}.txt"
        reply_path = os.path.join(outbox_dir, reply_filename)
        
        with open(reply_path, 'w', encoding='utf-8') as rf:
            rf.write(f"To: {customer_email}\n")
            rf.write(f"Subject: {reply_subject}\n")
            if reply_to_msg_id:
                rf.write(f"In-Reply-To: {reply_to_msg_id}\n")
            if pdf_path:
                rf.write(f"Attachment: {os.path.basename(pdf_path)}\n")
            rf.write("=" * 80 + "\n")
            rf.write(plain_body)
        print(f"[Mock Mailer] Written mock reply for tenant {tenant_id} to {reply_path}")
        
        if pdf_path and os.path.exists(pdf_path):
            outbox_pdf_path = os.path.join(outbox_dir, os.path.basename(pdf_path))
            if os.path.abspath(pdf_path) != os.path.abspath(outbox_pdf_path):
                import shutil
                try:
                    shutil.copy2(pdf_path, outbox_pdf_path)
                except Exception as ce:
                    print(f"[Mock Mailer] Warning: Failed to copy PDF to outbox: {ce}")
        return True
    else:
        try:
            reply_msg = MIMEMultipart()
            reply_msg["From"] = email_user
            reply_msg["To"] = customer_email
            reply_msg["Subject"] = reply_subject
            if reply_to_msg_id:

                reply_msg["In-Reply-To"] = reply_to_msg_id
                reply_msg["References"] = reply_to_msg_id

            msg_alt = MIMEMultipart('alternative')
            msg_alt.attach(MIMEText(plain_body, 'plain'))
            msg_alt.attach(MIMEText(html_body, 'html'))
            reply_msg.attach(msg_alt)
            
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            logo_file_path = tenant_config.get("company_logo_path")
            if logo_file_path:
                if not os.path.isabs(logo_file_path):
                    logo_file_path = os.path.join(project_root, logo_file_path)
            else:
                logo_file_path = find_company_logo(project_root)
                
            if logo_file_path and os.path.exists(logo_file_path):
                try:
                    from email.mime.image import MIMEImage
                    with open(logo_file_path, 'rb') as f:
                        logo_img = MIMEImage(f.read())
                    logo_img.add_header('Content-ID', '<company_logo>')
                    logo_img.add_header('Content-Disposition', 'inline', filename=os.path.basename(logo_file_path))
                    reply_msg.attach(logo_img)
                except Exception as le:
                    print(f"[Warning] Failed to attach inline logo: {le}")
            
            if pdf_path and os.path.exists(pdf_path):
                with open(pdf_path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(pdf_path)}")
                reply_msg.attach(part)
            
            if smtp_port == 465:
                server = smtplib.SMTP_SSL(smtp_server, smtp_port)
            else:
                server = smtplib.SMTP(smtp_server, smtp_port)
                server.starttls()
            server.login(email_user, email_pass)
            server.sendmail(email_user, customer_email, reply_msg.as_string())
            server.close()
            print(f"[SMTP Mailer] Successfully sent email to {customer_email} (Subject: {reply_subject})")
            return True
        except Exception as e:
            print(f"[SMTP Mailer] Error sending email via SMTP: {e}")
            return False

def poll_email_inbox(catalog, crm_path, mode="mock", tenant_id=None):
    """
    Main polling function. Checks for incoming mails (Live or Mock simulation).
    """
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    from src.tenants import get_tenant_config
    tenant_config = get_tenant_config(tenant_id)
    
    # Dynamically resolve CRM path from tenant config if configured
    if tenant_config and tenant_config.get("crm_json"):
        tenant_crm = tenant_config.get("crm_json")
        if not os.path.isabs(tenant_crm):
            crm_path = os.path.join(project_root, tenant_crm)
        else:
            crm_path = tenant_crm
        if not os.path.exists(crm_path):
            crm_path = os.path.join(project_root, "data", "crm_customers.json")
            
    crm_emails = load_crm_emails(crm_path)
    
    if mode == "live" and tenant_config and tenant_config.get("outlook_client_secret"):
        poll_outlook_graph(catalog, crm_path, tenant_id, tenant_config, crm_emails, project_root)
        return
        
    if mode == "mock":
        try:
            from src.database_sqlite import update_service_status
            update_service_status("CONNECTED", tenant_id=tenant_id)
        except Exception:
            pass
        if tenant_id and tenant_id != "default":
            inbox_dir = os.path.join(project_root, "mock_inbox", tenant_id)
            outbox_dir = os.path.join(project_root, "mock_outbox", tenant_id)
        else:
            inbox_dir = os.path.join(project_root, "mock_inbox")
            outbox_dir = os.path.join(project_root, "mock_outbox")
        os.makedirs(inbox_dir, exist_ok=True)
        os.makedirs(outbox_dir, exist_ok=True)
        
        files = [f for f in os.listdir(inbox_dir) if f.endswith(".txt")]
        if not files:
            try:
                from src.database_sqlite import update_service_status
                update_service_status("IDLE", tenant_id=tenant_id)
            except Exception:
                pass
            return
            
        print(f"[Email Listener] Found {len(files)} new enquiry files in mock_inbox/{tenant_id or ''}.")
        for file in files:
            file_path = os.path.join(inbox_dir, file)
            try:
                sender, subject, body, msg_id = parse_mock_email(file_path)
                
                # Extract details for notifications & reply detection first
                from email.utils import parseaddr
                display_name, email_addr = parseaddr(sender)
                _, crm_name, crm_phone = get_crm_discount(email_addr if email_addr else sender, crm_path)
                sender_name = crm_name
                if sender_name == "Walk-in Retail Client":
                    sender_name = display_name if display_name else (email_addr if email_addr else sender).split('@')[0]
                contact_phone = extract_phone_number(body)
                if not contact_phone:
                    contact_phone = crm_phone
                if not contact_phone:
                    contact_phone = "—"
                
                # ── Customer Reply Detection ─────────────────────────────────────────
                subj_lower = subject.lower().strip()
                is_reply_subject = (
                    subj_lower.startswith("re:") or
                    subj_lower.startswith("fw:") or
                    subj_lower.startswith("fwd:") or
                    "reply" in subj_lower or
                    "status" in subj_lower or
                    "quote" in subj_lower or
                    "follow" in subj_lower or
                    bool(re.search(r'\[Quotation\s+#|QTN-[A-Z0-9\-]+', subj_lower, re.IGNORECASE))
                )
                
                # Match alphanumeric quotation ID inside brackets [Quotation #ID] first, then fall back to QTN-XXXX
                qtn_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', subject, re.IGNORECASE)
                qtn_ref_extracted = None
                if qtn_match:
                    qtn_ref_extracted = qtn_match.group(1).upper()
                else:
                    qtn_match = re.search(r'(QTN-[A-Z0-9\-]+)', subject, re.IGNORECASE)
                    if qtn_match:
                        qtn_ref_extracted = qtn_match.group(1).upper()

                # Scan body for QTN reference
                if not qtn_ref_extracted:
                    body_qtn_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', body[:3000], re.IGNORECASE)
                    if body_qtn_match:
                        qtn_ref_extracted = body_qtn_match.group(1).upper()
                    else:
                        body_qtn_match2 = re.search(r'(QTN-[A-Z0-9\-]+)', body[:3000], re.IGNORECASE)
                        if body_qtn_match2:
                            qtn_ref_extracted = body_qtn_match2.group(1).upper()

                sender_qtn = None
                sender_email_normalized = (email_addr or sender).lower().strip() if (email_addr or sender) else ""
                if not qtn_ref_extracted and sender_email_normalized:
                    try:
                        from src.database_sqlite import get_connection
                        import datetime as _dt
                        _rc = get_connection(tenant_id)
                        _30d_ago = (_dt.datetime.utcnow() - _dt.timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
                        _row = _rc.execute(
                            "SELECT invoice_id FROM quotations WHERE LOWER(customer_email) = ? AND created_at >= ? ORDER BY rowid DESC LIMIT 1",
                            (sender_email_normalized, _30d_ago)
                        ).fetchone()
                        _rc.close()
                        if _row:
                            sender_qtn = _row[0]
                    except Exception as _e:
                        print(f"[Reply Detector] sender-quote lookup failed: {_e}")

                is_customer_reply = False
                qtn_ref = None
                candidate_qtn = qtn_ref_extracted or sender_qtn
                if candidate_qtn and (is_reply_subject or qtn_ref_extracted):
                    try:
                        from src.database_sqlite import get_connection
                        _rc = get_connection(tenant_id)
                        _exists = _rc.execute("SELECT 1 FROM quotations WHERE invoice_id = ?", (candidate_qtn,)).fetchone()
                        _rc.close()
                        if _exists:
                            is_customer_reply = True
                            qtn_ref = candidate_qtn
                    except Exception as _e:
                        print(f"[Reply Detector] quote existence check failed: {_e}")
                
                _reply_customer_msg_logged = False

                # Tier 1: Fast blocklist check (0ms, no API)
                blocklist_result = fast_blocklist_check(sender, subject, crm_emails)
                if blocklist_result == "REJECT":
                    print(f"[Email Filter] Skipped irrelevant mock email from {sender} (Subject: {subject})")
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    continue
                
                # Tier 2: AI relevance classification
                needs_relevance_check = (blocklist_result == "NEEDS_AI") or (blocklist_result == "ACCEPT_CRM" and not is_customer_reply)
                if needs_relevance_check:
                    # Fast relevance check on subject before using AI API
                    if not is_customer_reply and not is_subject_relevant(subject, sender, crm_emails, tenant_id=tenant_id):
                        print(f"[Email Filter] Skipped irrelevant mock email from {sender} (Subject: {subject}) [Fast Subject Check]")
                        if os.path.exists(file_path):
                            os.remove(file_path)
                        continue
                    ai_result = classify_and_extract(sender, subject, body)
                    if ai_result and ai_result.get("intent") == "IRRELEVANT":
                        print(f"[AI Filter] Skipped irrelevant mock email from {sender} (Subject: {subject}) — Confidence: {ai_result.get('confidence', 0):.2f}")
                        if os.path.exists(file_path):
                            os.remove(file_path)
                        continue
                    elif ai_result is None:
                        # API unavailable — fall back to existing rule-based filter
                        if not is_email_relevant(sender, subject, body, catalog, crm_emails):
                            print(f"[Email Filter] Skipped irrelevant mock email from {sender} (Subject: {subject}) [rule-based fallback]")
                            if os.path.exists(file_path):
                                os.remove(file_path)
                            continue
                
                print(f"\n[Processing Mock Email] From: {sender} | Subject: {subject}")
                
                # Fetch credentials from tenant config
                master_email = tenant_config.get("master_email")
                email_user = tenant_config.get("email_user")
                email_pass = tenant_config.get("email_pass")
                smtp_server = tenant_config.get("smtp_server", "smtp.gmail.com")
                try:
                    smtp_port = int(tenant_config.get("smtp_port", 465))
                except (ValueError, TypeError):
                    smtp_port = 465

                # ── Immediate DB Logging (Stage tracking) ──────────────────────────────
                import datetime
                tz_ist = datetime.timezone(datetime.timedelta(hours=5, minutes=30))
                received_at = datetime.datetime.now(tz_ist).strftime("%Y-%m-%d %H:%M:%S")
                if msg_id:
                    from src.database_sqlite import log_processed_message
                    if is_customer_reply:
                        log_processed_message(msg_id, f"CUSTOMER_REPLIED:{qtn_ref}", received_at=received_at, customer_name=sender_name, customer_email=(email_addr or sender), tenant_id=tenant_id)
                    else:
                        log_processed_message(msg_id, "NEW", received_at=received_at, customer_name=sender_name, customer_email=(email_addr or sender), tenant_id=tenant_id)

                # Log timeline comment if customer reply
                if is_customer_reply and qtn_ref and qtn_ref != "REPLIED":
                    try:
                        from src.database_sqlite import log_chat_msg
                        _reply = body or ""
                        _cut = re.search(r'On\s+.+?\bwrote:', _reply, re.DOTALL)
                        if _cut:
                            _reply = _reply[:_cut.start()]
                        _reply_clean = _reply.strip() or "(customer replied)"
                        log_chat_msg(qtn_ref, "customer", _reply_clean, tenant_id=tenant_id)
                        _reply_customer_msg_logged = True
                    except Exception as _e:
                        print(f"[Reply Detector] failed to log reply to timeline: {_e}")

                # 1. Notify Master User of incoming enquiry
                if master_email and email_user and email_pass and email_user.strip() and not email_user.startswith("your_"):
                    subject_notif = f"[Notification] New enquiry received from {sender_name}"
                    body_notif = (
                        f"Dear Master User,\n\n"
                        f"An enquiry email has been received and processed in mock mode.\n\n"
                        f"Customer Details:\n"
                        f"- Name: {sender_name}\n"
                        f"- Email: {email_addr if email_addr else sender}\n"
                        f"- Contact Number: {contact_phone}\n\n"
                        f"Original Message Subject: {subject}\n"
                        f"Original Message Body:\n"
                        f"--------------------------------------------------\n"
                        f"{body}\n"
                        f"--------------------------------------------------\n\n"
                        f"Regards,\n"
                        f"Trofeo Auto-bot"
                    )
                    send_master_notification(smtp_server, smtp_port, email_user, email_pass, master_email, subject_notif, body_notif)

                # ── Human Request Checking & Routing ───────────────────────────────
                is_human = is_human_request(body)
                if is_human:
                    print(f"[Human Request] Customer requested human agent. Routing to Pending.")
                    reply_subject = clean_reply_subject(subject, is_unparsed=True)
                    reply_body = (
                        f"Dear {sender_name},\n\n"
                        f"Thank you for contacting us. We have received your request and forwarded it "
                        f"to a sales representative. A team member will contact you shortly to assist you.\n\n"
                        f"Regards,\n"
                        f"{tenant_config.get('business_name', 'Trofeo Hardware')} Sales Team"
                    )
                    plain_body = reply_body
                    pdf_path = None
                    status = "UNPARSED_NOTICE"
                    proc_invoice_id = None
                else:
                    result = process_incoming_email(
                        sender, subject, body, catalog, crm_path, mode, project_root, tenant_id=tenant_id,
                        skip_initial_customer_log=_reply_customer_msg_logged
                    )
                    if len(result) == 5:
                        reply_subject, reply_body_tuple, pdf_path, status, proc_invoice_id = result
                    else:
                        reply_subject, reply_body_tuple, pdf_path, status = result
                        proc_invoice_id = getattr(result, "invoice_id", None)
                    
                    if not proc_invoice_id:
                        quote_id_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', reply_subject, re.IGNORECASE)
                        if quote_id_match:
                            proc_invoice_id = quote_id_match.group(1)
                    
                    if isinstance(reply_body_tuple, tuple):
                        plain_body = reply_body_tuple[0]
                    else:
                        plain_body = reply_body_tuple
                
                # 2. Notify Master User of outgoing reply
                if master_email and email_user and email_pass and email_user.strip() and not email_user.startswith("your_"):
                    subject_reply_notif = f"[Notification] Reply sent to {sender_name}"
                    body_reply_notif = (
                        f"Dear Master User,\n\n"
                        f"The bot has successfully sent a reply to the customer's enquiry in mock mode.\n\n"
                        f"Customer Details:\n"
                        f"- Name: {sender_name}\n"
                        f"- Email: {email_addr if email_addr else sender}\n"
                        f"- Contact Number: {contact_phone}\n\n"
                        f"Reply Details:\n"
                        f"- Subject: {reply_subject}\n"
                        f"- Status: {status}\n\n"
                        f"Replied Message Content:\n"
                        f"--------------------------------------------------\n"
                        f"{plain_body}\n"
                        f"--------------------------------------------------\n\n"
                        f"Regards,\n"
                        f"Trofeo Auto-bot"
                    )
                    send_master_notification(smtp_server, smtp_port, email_user, email_pass, master_email, subject_reply_notif, body_reply_notif)

                # --- Handle mixed matched/unmatched items (customer asks for products not in the catalog) ---
                if status in ["QUOTE_GENERATED", "QUOTE_UPDATED"]:
                    clean_body_for_deficits = strip_email_history(body)
                    has_attachment = "[From attachment" in clean_body_for_deficits
                    cap_by_stock = False
                    override_qty = None
                    if has_attachment:
                        cap_by_stock = True
                        parts = clean_body_for_deficits.split("[From attachment")
                        email_text_only = parts[0].strip()
                        override_qty = extract_global_quantity_override(email_text_only)
                        body_lines = run_scenario_free(email_text_only, catalog)
                        body_has_products = any(l["matched_sku_id"] != "UNKNOWN" for l in body_lines)
                        if body_has_products or override_qty is not None:
                            cap_by_stock = False

                    matched_lines = run_scenario_free(clean_body_for_deficits, catalog)
                    if override_qty is not None:
                        for line in matched_lines:
                            if line['matched_sku_id'] != "UNKNOWN":
                                line['quantity'] = override_qty
                    deficit_lines = adjust_quantities_by_stock(matched_lines, catalog, cap_by_stock=cap_by_stock)
                    
                    # Trigger deficit purchase order alert if there are deficits
                    if deficit_lines:
                        print(f"[Deficit PO] {len(deficit_lines)} deficit items found. Sending alert to master.")
                        if master_email and email_user and email_pass and email_user.strip() and not email_user.startswith("your_"):
                            send_deficit_purchase_order_alert(
                                smtp_server=smtp_server,
                                smtp_port=smtp_port,
                                email_user=email_user,
                                email_pass=email_pass,
                                master_email=master_email,
                                customer_name=sender_name,
                                customer_email=email_addr if email_addr else sender,
                                customer_phone=contact_phone,
                                original_subject=subject,
                                deficit_lines=deficit_lines
                            )
                            
                    unmatched_lines = [line for line in matched_lines if line['matched_sku_id'] == "UNKNOWN"]
                    if unmatched_lines:
                        print(f"[Unmatched Products] {len(unmatched_lines)} unmatched items found. Logging & alerting master.")
                        try:
                            from src.database_sqlite import log_unmatched_item
                            unmatched_desc = "\n".join([f"{l['quantity']} x {l['parsed_query']} (original: {l['original_line']})" for l in unmatched_lines])
                            log_unmatched_item(
                                customer_email=sender,
                                customer_name=sender_name,
                                original_body=f"UNMATCHED PRODUCTS REQUESTED:\n{unmatched_desc}\n\nFULL EMAIL BODY:\n{body}",
                                source="mock_email",
                                tenant_id=tenant_id
                            )
                        except Exception as ue:
                            print(f"[Warning] Failed to log unmatched items: {ue}")

                        if master_email and email_user and email_pass and email_user.strip() and not email_user.startswith("your_"):
                            send_unmatched_products_alert(
                                smtp_server=smtp_server,
                                smtp_port=smtp_port,
                                email_user=email_user,
                                email_pass=email_pass,
                                master_email=master_email,
                                customer_name=sender_name,
                                customer_email=email_addr if email_addr else sender,
                                customer_phone=contact_phone,
                                original_subject=subject,
                                unmatched_lines=unmatched_lines
                            )

                # --- Handle UNPARSED_NOTICE: log unmatched items + alert master ---
                if status == "UNPARSED_NOTICE":
                    print(f"[Unmatched] No valid SKU matches for mock enquiry from {sender}. Logging & alerting master (No reply generated).")
                    try:
                        from src.database_sqlite import log_unmatched_item
                        log_unmatched_item(
                            customer_email=sender,
                            customer_name=sender_name,
                            original_body=body,
                            source="mock_email",
                            tenant_id=tenant_id
                        )
                    except Exception as ue:
                        print(f"[Warning] Failed to log unmatched item: {ue}")

                    if master_email and email_user and email_pass and email_user.strip() and not email_user.startswith("your_"):
                        subject_unmatch_notif = f"[ACTION REQUIRED] Unmatched enquiry from {sender_name} — Manual Quote Needed"
                        body_unmatch_notif = (
                            f"Dear Master User,\n\n"
                            f"The auto-bot received an enquiry but could NOT prepare an automated quote because "
                            f"the requested items do not match any product in our catalogue.\n\n"
                            f"Customer Details:\n"
                            f"- Name: {sender_name}\n"
                            f"- Email: {email_addr if email_addr else sender}\n"
                            f"- Contact Number: {contact_phone}\n\n"
                            f"Original Enquiry:\n"
                            f"Subject: {subject}\n"
                            f"--------------------------------------------------\n"
                            f"{body}\n"
                            f"--------------------------------------------------\n\n"
                            f"Please review the above request and prepare a manual quotation directly to the customer.\n\n"
                            f"This enquiry has been logged in the database for your reference.\n\n"
                            f"Regards,\n"
                            f"Trofeo Auto-bot"
                        )
                        send_master_notification(smtp_server, smtp_port, email_user, email_pass, master_email, subject_unmatch_notif, body_unmatch_notif)
                    
                    # Log processed message reference, log activity, and skip writing reply file
                    log_invoice_ref = "UNMATCHED"
                    if msg_id:
                        from src.database_sqlite import log_processed_message
                        try:
                            from src.database_sqlite import get_connection
                            _rc = get_connection(tenant_id)
                            _row = _rc.execute(
                                "SELECT id FROM unmatched_items WHERE customer_email = ? ORDER BY id DESC LIMIT 1",
                                (sender,)
                            ).fetchone()
                            log_invoice_ref = f"UNMATCHED_{_row[0]}" if _row else "UNMATCHED"
                        except Exception:
                            log_invoice_ref = "UNMATCHED"
                        log_processed_message(msg_id, log_invoice_ref, received_at=received_at, tenant_id=tenant_id)
                    
                    try:
                        from src.database_sqlite import log_activity
                        log_activity("UNMATCHED_ENQUIRY", invoice_id=log_invoice_ref,
                            customer_name=sender_name, customer_email=sender,
                            description=f"Items not found in catalog (Subject: {subject[:80]})",
                            tenant_id=tenant_id)
                    except Exception:
                        pass
                    continue

                # Check reply mode setting
                from src.database_sqlite import get_setting, update_quotation_status, log_chat_msg
                reply_mode = get_setting("reply_mode", "auto", tenant_id)
                
                is_held_draft = (status in ("QUOTE_GENERATED", "QUOTE_UPDATED", "CONVERSATIONAL_REPLY", "CUSTOMER_REPLIED") or is_customer_reply)
                
                reply_filename = file.replace(".txt", "_reply.txt")
                reply_path = os.path.join(outbox_dir, reply_filename)
                
                sent = False
                if reply_mode == "manual" and is_held_draft:
                    ref_id = proc_invoice_id or qtn_ref
                    if ref_id:
                        update_quotation_status(ref_id, "PENDING_REVIEW", tenant_id=tenant_id)
                    
                    draft_msg = f"Subject: {reply_subject}\n\n{plain_body}"
                    log_chat_msg(ref_id or "MANUAL", "DRAFT_BOT", draft_msg, tenant_id=tenant_id)
                    print(f"[Mock Mailer] Manual mode active. Holding reply draft {ref_id} for approval.")
                    sent = True
                else:
                    try:
                        with open(reply_path, 'w', encoding='utf-8') as rf:
                            rf.write(f"To: {sender}\n")
                            rf.write(f"Subject: {reply_subject}\n")
                            if msg_id:
                                rf.write(f"In-Reply-To: {msg_id}\n")
                                rf.write(f"References: {msg_id}\n")
                            if pdf_path:
                                rf.write(f"Attachment: {os.path.basename(pdf_path)}\n")
                            rf.write("=" * 80 + "\n")
                            rf.write(plain_body)
                        sent = True
                    except Exception as e:
                        print(f"[Mock Mailer] Error writing mock reply file: {e}")
                
                if sent:
                    if msg_id:
                        if not (reply_mode == "manual" and is_held_draft):
                            if is_customer_reply and qtn_ref and qtn_ref != "REPLIED" and plain_body:
                                try:
                                    log_chat_msg(qtn_ref, "BOT", plain_body.strip(), tenant_id=tenant_id)
                                except Exception as _ce:
                                    print(f"[Reply Detector] failed to log bot reply to timeline: {_ce}")
                        from src.database_sqlite import log_processed_message
                        if is_customer_reply:
                            log_invoice_ref = f"CUSTOMER_REPLIED:{qtn_ref}"
                        elif status == "UNPARSED_NOTICE":
                            try:
                                from src.database_sqlite import get_connection
                                _rc = get_connection(tenant_id)
                                _row = _rc.execute(
                                    "SELECT id FROM unmatched_items WHERE customer_email = ? ORDER BY id DESC LIMIT 1",
                                    (sender,)
                                ).fetchone()
                                log_invoice_ref = f"UNMATCHED_{_row[0]}" if _row else "UNMATCHED"
                            except Exception:
                                log_invoice_ref = "UNMATCHED"
                        else:
                            log_invoice_ref = proc_invoice_id or status
                        log_processed_message(msg_id, log_invoice_ref, received_at=received_at, tenant_id=tenant_id)
                        
                        # Log EMAIL_SENT activity (Bug 12)
                        try:
                            from src.database_sqlite import log_activity
                            _sent_inv = proc_invoice_id or (f"CUSTOMER_REPLIED:{qtn_ref}" if is_customer_reply and qtn_ref else None)
                            log_activity("EMAIL_SENT", invoice_id=_sent_inv,
                                customer_name=sender_name, customer_email=sender,
                                description=f"Reply sent to {sender} — Subject: {reply_subject[:80]}",
                                tenant_id=tenant_id)
                        except Exception:
                            pass

                print(f"[Success] Processed (status: {status}) for tenant {tenant_id}. Written reply & quote to mock_outbox/.")
                
            except Exception as e:
                print(f"[Error] Failed to process mock email {file} for tenant {tenant_id}: {e}")
            finally:
                if os.path.exists(file_path):
                    os.remove(file_path)
        try:
            from src.database_sqlite import update_service_status
            update_service_status("IDLE", tenant_id=tenant_id)
        except Exception:
            pass
            
    else:
        # Live IMAP/SMTP Pipeline
        imap_server = tenant_config.get("imap_server", "imap.gmail.com")
        try:
            imap_port = int(tenant_config.get("imap_port", 993))
        except (ValueError, TypeError):
            imap_port = 993
            
        smtp_server = tenant_config.get("smtp_server", "smtp.gmail.com")
        try:
            smtp_port = int(tenant_config.get("smtp_port", 465))
        except (ValueError, TypeError):
            smtp_port = 465
            
        email_user = tenant_config.get("email_user")
        email_pass = tenant_config.get("email_pass")
        
        if not email_user or not email_pass:
            print(f"[Email Listener] Error: Credentials missing in tenant {tenant_id} config.")
            return
  
        try:
            mail = imaplib.IMAP4_SSL(imap_server, imap_port)
            mail.login(email_user, email_pass)
            try:
                from src.database_sqlite import update_service_status
                update_service_status("CONNECTED", tenant_id=tenant_id)
            except Exception:
                pass
            mail.select("inbox")
            
            status, messages = mail.search(None, 'ALL')
            if status == "OK" and messages[0]:
                mail_ids = messages[0].split()
                mail_ids = list(reversed(mail_ids))[:30]
                print(f"[Email Listener] Found {len(messages[0].split())} total emails in inbox for {tenant_id}. Checking the {len(mail_ids)} newest.")
                
                for m_id in mail_ids:
                    res, msg_data = mail.fetch(m_id, '(BODY[HEADER.FIELDS (SUBJECT FROM MESSAGE-ID DATE)])')
                    if res != "OK" or not msg_data or not msg_data[0]:
                        continue
                    
                    raw_headers = msg_data[0][1]
                    msg_headers = email.message_from_bytes(raw_headers)
                    
                    subject_raw = msg_headers.get("Subject", "Order Enquiry")
                    
                    subject = ""
                    try:
                        parts = email.header.decode_header(subject_raw)
                        for part, encoding in parts:
                            if isinstance(part, bytes):
                                subject += part.decode(encoding or 'utf-8', errors='ignore')
                            else:
                                subject += part
                    except Exception:
                        subject = str(subject_raw)
                    
                    sender_header = msg_headers.get("From", "")
                    sender = email.utils.parseaddr(sender_header)[1]
                    msg_id = msg_headers.get("Message-ID")
                    date_raw = msg_headers.get("Date")
                    received_at = format_email_date(date_raw)
                    
                    # Skip if this specific email message has already been processed (based on unique Message-ID)
                    if msg_id:
                        from src.database_sqlite import is_message_processed
                        if is_message_processed(msg_id, tenant_id=tenant_id):
                            continue
                            
                    # Prevent infinite loop by skipping emails sent by ourselves, unless it's a test enquiry (subject doesn't start with '[')
                    if email_user and sender.lower() == email_user.lower():
                        subj_clean = subject.strip()
                        if subj_clean.startswith('[') or 'quotation' in subj_clean.lower() or 'notification' in subj_clean.lower() or 'action required' in subj_clean.lower():
                            print(f"[Email Listener] Ignored email from ourselves ({sender}) with subject '{subject}' to prevent loop.")
                            mail.store(m_id, '+FLAGS', '\\Seen')
                            if msg_id:
                                from src.database_sqlite import log_processed_message
                                log_processed_message(msg_id, "SELF_SENT", received_at=received_at, tenant_id=tenant_id)
                            continue
                        else:
                            print(f"[Email Listener] Processing self-sent test email: Subject='{subject}'")
                    
                    res, msg_data = mail.fetch(m_id, '(RFC822)')
                    if res != "OK" or not msg_data or not msg_data[0]:
                        continue
                        
                    raw_email = msg_data[0][1]
                    msg = email.message_from_bytes(raw_email)
                    
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                                break
                    else:
                        body = msg.get_payload(decode=True).decode('utf-8', errors='ignore')

                    # Quick check: does this email have any attachments?
                    email_has_attach = has_attachments(msg)

                    # Extract text from attachments first to include it in the classification and parsing
                    attachment_text = ""
                    clean_attach_text = ""
                    if email_has_attach:
                        attachment_text = extract_text_from_attachments(msg)
                        clean_attach_text = "" if not attachment_text else "\n".join(
                            line for line in attachment_text.splitlines()
                            if not line.startswith("[ATTACHMENT_FAILED:")
                        ).strip()
                        if clean_attach_text:
                            print(f"[Attachment] Successfully extracted text from attachments. Merging with body.")
                            body = (body + "\n\n" + clean_attach_text).strip()

                    from email.utils import parseaddr
                    display_name, email_addr = parseaddr(sender_header)
                    _, crm_name, crm_phone = get_crm_discount(sender, crm_path)
                    sender_name = crm_name
                    if sender_name == "Walk-in Retail Client":
                        sender_name = display_name if display_name else sender.split('@')[0]
                    contact_phone = extract_phone_number(body)
                    if not contact_phone:
                        contact_phone = crm_phone
                    if not contact_phone:
                        contact_phone = "—"
                        
                    # ── Customer Reply Detection ─────────────────────────────────────────
                    subj_lower = subject.lower().strip()
                    is_reply_subject = (
                        subj_lower.startswith("re:") or
                        subj_lower.startswith("fw:") or
                        subj_lower.startswith("fwd:") or
                        "reply" in subj_lower or
                        "status" in subj_lower or
                        "quote" in subj_lower or
                        "follow" in subj_lower or
                        bool(re.search(r'\[Quotation\s+#|QTN-[A-Z0-9\-]+', subj_lower, re.IGNORECASE))
                    )
                    
                    # Match alphanumeric quotation ID inside brackets [Quotation #ID] first, then fall back to QTN-XXXX
                    qtn_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', subject, re.IGNORECASE)
                    qtn_ref_extracted = None
                    if qtn_match:
                        qtn_ref_extracted = qtn_match.group(1).upper()
                    else:
                        qtn_match = re.search(r'(QTN-[A-Z0-9\-]+)', subject, re.IGNORECASE)
                        if qtn_match:
                            qtn_ref_extracted = qtn_match.group(1).upper()

                    # Scan body for QTN reference
                    if not qtn_ref_extracted:
                        body_qtn_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', body[:3000], re.IGNORECASE)
                        if body_qtn_match:
                            qtn_ref_extracted = body_qtn_match.group(1).upper()
                        else:
                            body_qtn_match2 = re.search(r'(QTN-[A-Z0-9\-]+)', body[:3000], re.IGNORECASE)
                            if body_qtn_match2:
                                qtn_ref_extracted = body_qtn_match2.group(1).upper()

                    sender_qtn = None
                    sender_email_normalized = sender.lower().strip() if sender else ""
                    if not qtn_ref_extracted and sender_email_normalized:
                        try:
                            from src.database_sqlite import get_connection
                            import datetime as _dt
                            _rc = get_connection(tenant_id)
                            _30d_ago = (_dt.datetime.utcnow() - _dt.timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
                            _row = _rc.execute(
                                "SELECT invoice_id FROM quotations WHERE LOWER(customer_email) = ? AND created_at >= ? ORDER BY rowid DESC LIMIT 1",
                                (sender_email_normalized, _30d_ago)
                            ).fetchone()
                            _rc.close()
                            if _row:
                                sender_qtn = _row[0]
                        except Exception as _e:
                            print(f"[Reply Detector] sender-quote lookup failed: {_e}")

                    is_customer_reply = False
                    qtn_ref = None
                    candidate_qtn = qtn_ref_extracted or sender_qtn
                    if candidate_qtn and (is_reply_subject or qtn_ref_extracted):
                        try:
                            from src.database_sqlite import get_connection
                            _rc = get_connection(tenant_id)
                            _exists = _rc.execute("SELECT 1 FROM quotations WHERE invoice_id = ?", (candidate_qtn,)).fetchone()
                            _rc.close()
                            if _exists:
                                is_customer_reply = True
                                qtn_ref = candidate_qtn
                        except Exception as _e:
                            print(f"[Reply Detector] quote existence check failed: {_e}")
                    
                    _reply_customer_msg_logged = False

                    # Tier 1: Fast blocklist check (0ms, no API)
                    blocklist_result = fast_blocklist_check(sender, subject, crm_emails)
                    if blocklist_result == "REJECT":
                        print(f"[Email Filter] Skipped irrelevant email from {sender} (Subject: {subject})")
                        mail.store(m_id, '+FLAGS', '\\Seen')
                        if msg_id:
                            from src.database_sqlite import log_processed_message
                            log_processed_message(msg_id, "IRRELEVANT", received_at=received_at, tenant_id=tenant_id)
                        continue
                    
                    # Tier 2: AI relevance classification
                    needs_relevance_check = (blocklist_result == "NEEDS_AI") or (blocklist_result == "ACCEPT_CRM" and not is_customer_reply)
                    if needs_relevance_check:
                        # Fast relevance check on subject before using AI API
                        if not is_customer_reply and not is_subject_relevant(subject, sender, crm_emails, tenant_id=tenant_id):
                            print(f"[Email Filter] Skipped irrelevant email subject from {sender} (Subject: {subject}) [Fast Subject Check]")
                            mail.store(m_id, '+FLAGS', '\\Seen')
                            if msg_id:
                                from src.database_sqlite import log_processed_message
                                log_processed_message(msg_id, "IRRELEVANT", received_at=received_at, tenant_id=tenant_id)
                            continue
                        ai_result = classify_and_extract(sender, subject, body)
                        if ai_result and ai_result.get("intent") == "IRRELEVANT":
                            print(f"[AI Filter] Skipped irrelevant email from {sender} (Subject: {subject}) — Confidence: {ai_result.get('confidence', 0):.2f}")
                            mail.store(m_id, '+FLAGS', '\\Seen')
                            if msg_id:
                                from src.database_sqlite import log_processed_message
                                log_processed_message(msg_id, "IRRELEVANT", received_at=received_at, tenant_id=tenant_id)
                            continue
                        elif ai_result is None:
                            # API unavailable — fall back to existing rule-based filter
                            if not is_email_relevant(sender, subject, body, catalog, crm_emails,
                                                     attachment_text=clean_attach_text, email_has_attachments=email_has_attach):
                                print(f"[Email Filter] Skipped irrelevant email from {sender} (Subject: {subject}) [rule-based fallback]")
                                mail.store(m_id, '+FLAGS', '\\Seen')
                                if msg_id:
                                    from src.database_sqlite import log_processed_message
                                    log_processed_message(msg_id, "IRRELEVANT", received_at=received_at, tenant_id=tenant_id)
                                continue

                    if not body.strip():
                        print(f"[Email Filter] Email from {sender} has no parseable text content.")
                        master_email_addr = tenant_config.get("master_email")
                        if master_email_addr and email_has_attach:
                            from email.utils import parseaddr as _parseaddr
                            _dname, _eaddr = _parseaddr(sender_header)
                            _sname = _dname if _dname else (_eaddr or sender).split('@')[0]
                            _attach_note = "The email contained attachment(s), but automatic text extraction failed."
                            _notif_subj = f"[ACTION REQUIRED] Cannot auto-process attachment enquiry from {_sname}"
                            _notif_body = (
                                f"Dear Master User,\n\n"
                                f"An enquiry email was received from a customer, but the system could not "
                                f"extract readable product information from the attached file(s).\n\n"
                                f"Customer Details:\n"
                                f"- Name: {_sname}\n"
                                f"- Email: {sender}\n\n"
                                f"Email Subject: {subject}\n\n"
                                f"Note: {_attach_note}\n\n"
                                f"Regards,\n"
                                f"{tenant_config.get('business_name', 'Trofeo Hardware')} Auto-bot"
                            )
                            send_master_notification(
                                smtp_server, smtp_port, email_user, email_pass,
                                master_email_addr, _notif_subj, _notif_body
                            )
                        mail.store(m_id, '+FLAGS', '\\Seen')
                        if msg_id:
                            from src.database_sqlite import log_processed_message
                            log_processed_message(msg_id, "EMPTY_BODY", received_at=received_at, tenant_id=tenant_id)
                        continue

                    print(f"\n[Processing Live Email] From: {sender} | Subject: {subject} | Tenant: {tenant_id}")

                    # ── Immediate DB Logging (Stage tracking) ──────────────────────────────
                    if msg_id:
                        from src.database_sqlite import log_processed_message
                        if is_customer_reply:
                            log_processed_message(msg_id, f"CUSTOMER_REPLIED:{qtn_ref}", received_at=received_at, tenant_id=tenant_id)
                        else:
                            log_processed_message(msg_id, "NEW", received_at=received_at, tenant_id=tenant_id)

                    # Log timeline comment if customer reply
                    if is_customer_reply and qtn_ref and qtn_ref != "REPLIED":
                        try:
                            from src.database_sqlite import log_chat_msg
                            _reply = body or ""
                            _cut = re.search(r'On\s+.+?\bwrote:', _reply, re.DOTALL)
                            if _cut:
                                _reply = _reply[:_cut.start()]
                            _reply_clean = _reply.strip() or "(customer replied)"
                            log_chat_msg(qtn_ref, "customer", _reply_clean, tenant_id=tenant_id)
                            _reply_customer_msg_logged = True
                        except Exception as _e:
                            print(f"[Reply Detector] failed to log reply to timeline: {_e}")

                    # 1. Notify Master User of incoming enquiry
                    master_email = tenant_config.get("master_email")
                    if master_email and email_user and email_pass:
                        subject_notif = f"[Notification] New enquiry received from {sender_name}"
                        body_notif = (
                            f"Dear Master User,\n\n"
                            f"An enquiry email has been received and processed.\n\n"
                            f"Customer Details:\n"
                            f"- Name: {sender_name}\n"
                            f"- Email: {sender}\n"
                            f"- Contact Number: {contact_phone}\n\n"
                            f"Original Message Subject: {subject}\n"
                            f"Original Message Body:\n"
                            f"--------------------------------------------------\n"
                            f"{body}\n"
                            f"--------------------------------------------------\n\n"
                            f"Regards,\n"
                            f"Trofeo Auto-bot"
                        )
                        send_master_notification(smtp_server, smtp_port, email_user, email_pass, master_email, subject_notif, body_notif)
                    
                    # ── Human Request Checking & Routing ───────────────────────────────
                    is_human = is_human_request(body)
                    if is_human:
                        print(f"[Human Request] Customer requested human agent. Routing to Pending.")
                        try:
                            from src.database_sqlite import log_unmatched_item
                            u_id = log_unmatched_item(
                                customer_email=sender,
                                customer_name=sender_name,
                                original_body=f"HUMAN AGENT REQUESTED:\n{body}",
                                source="live_email",
                                tenant_id=tenant_id
                            )
                            proc_invoice_id = f"UNMATCHED_{u_id}"
                        except Exception as ue:
                            print(f"[Warning] Failed to log human request unmatched item: {ue}")
                            proc_invoice_id = "UNMATCHED"
                        
                        reply_subject = clean_reply_subject(subject, is_unparsed=True)
                        reply_body = (
                            f"Dear {sender_name},\n\n"
                            f"Thank you for contacting us. We have received your request and forwarded it "
                            f"to a sales representative. A team member will contact you shortly to assist you.\n\n"
                            f"Regards,\n"
                            f"{tenant_config.get('business_name', 'Trofeo Hardware')} Sales Team"
                        )
                        plain_body = reply_body
                        html_body = f"<html><body><p>{reply_body.replace(chr(10), '<br>')}</p></body></html>"
                        pdf_path = None
                        status = "UNPARSED_NOTICE"
                    else:
                        result = process_incoming_email(
                            sender_header, subject, body, catalog, crm_path, mode, project_root, tenant_id=tenant_id,
                            skip_initial_customer_log=_reply_customer_msg_logged
                        )
                        if len(result) == 5:
                            reply_subject, reply_body_tuple, pdf_path, status, proc_invoice_id = result
                        else:
                            reply_subject, reply_body_tuple, pdf_path, status = result
                            proc_invoice_id = getattr(result, "invoice_id", None)
                        
                        if not proc_invoice_id:
                            quote_id_match = re.search(r'\[Quotation\s+#([A-Z0-9\-]+)\]', reply_subject, re.IGNORECASE)
                            if quote_id_match:
                                proc_invoice_id = quote_id_match.group(1)

                        if isinstance(reply_body_tuple, tuple):
                            plain_body, html_body = reply_body_tuple
                        else:
                            plain_body = reply_body_tuple
                            html_body = f"<html><body><p>{plain_body.replace('\n', '<br>')}</p></body></html>"
                    
                    reply_msg = MIMEMultipart()
                    reply_msg["From"] = email_user
                    reply_msg["To"] = sender
                    reply_msg["Subject"] = reply_subject
                    
                    if msg_id:
                        reply_msg["In-Reply-To"] = msg_id
                        reply_msg["References"] = msg_id
                        
                    msg_alt = MIMEMultipart('alternative')
                    msg_alt.attach(MIMEText(plain_body, 'plain'))
                    msg_alt.attach(MIMEText(html_body, 'html'))
                    reply_msg.attach(msg_alt)
                    
                    # Attach logo image inline
                    logo_file_path = tenant_config.get("company_logo_path")
                    if logo_file_path:
                        if not os.path.isabs(logo_file_path):
                            logo_file_path = os.path.join(project_root, logo_file_path)
                    else:
                        logo_file_path = find_company_logo(project_root)
                        
                    if logo_file_path and os.path.exists(logo_file_path):
                        try:
                            from email.mime.image import MIMEImage
                            with open(logo_file_path, 'rb') as f:
                                logo_img = MIMEImage(f.read())
                            logo_img.add_header('Content-ID', '<company_logo>')
                            logo_img.add_header('Content-Disposition', 'inline', filename=os.path.basename(logo_file_path))
                            reply_msg.attach(logo_img)
                        except Exception as le:
                            print(f"[Warning] Failed to attach inline logo: {le}")
                    
                    if pdf_path:
                        with open(pdf_path, "rb") as f:
                            part = MIMEBase("application", "octet-stream")
                            part.set_payload(f.read())
                        encoders.encode_base64(part)
                        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(pdf_path)}")
                        reply_msg.attach(part)
                    
                    # Check reply mode setting
                    from src.database_sqlite import get_setting, update_quotation_status, log_chat_msg
                    reply_mode = get_setting("reply_mode", "auto", tenant_id)
                    
                    is_held_draft = (status in ("QUOTE_GENERATED", "QUOTE_UPDATED", "CONVERSATIONAL_REPLY", "CUSTOMER_REPLIED") or is_customer_reply)
                    
                    sent = False
                    if reply_mode == "manual" and is_held_draft:
                        ref_id = proc_invoice_id or qtn_ref
                        if ref_id:
                            update_quotation_status(ref_id, "PENDING_REVIEW", tenant_id=tenant_id)
                        
                        draft_msg = f"Subject: {reply_subject}\n\n{plain_body}"
                        log_chat_msg(ref_id or "MANUAL", "DRAFT_BOT", draft_msg, tenant_id=tenant_id)
                        print(f"[SMTP Mailer] Manual mode active. Holding reply draft {ref_id} for approval.")
                        sent = True
                    else:
                        try:
                            # Send via SMTP
                            if smtp_port == 465:
                                server = smtplib.SMTP_SSL(smtp_server, smtp_port)
                            else:
                                server = smtplib.SMTP(smtp_server, smtp_port)
                                server.starttls()
                            server.login(email_user, email_pass)
                            server.sendmail(email_user, sender, reply_msg.as_string())
                            server.close()
                            sent = True
                        except Exception as e:
                            print(f"[SMTP Mailer] Error sending email via SMTP: {e}")
                    
                    if sent:
                        mail.store(m_id, '+FLAGS', '\\Seen')
                        if msg_id:
                            if not (reply_mode == "manual" and is_held_draft):
                                if is_customer_reply and qtn_ref and qtn_ref != "REPLIED" and plain_body:
                                    try:
                                        log_chat_msg(qtn_ref, "BOT", plain_body.strip(), tenant_id=tenant_id)
                                    except Exception as _ce:
                                        print(f"[Reply Detector] failed to log bot reply to timeline: {_ce}")
                            from src.database_sqlite import log_processed_message
                            if is_customer_reply:
                                log_invoice_ref = f"CUSTOMER_REPLIED:{qtn_ref}"
                            elif status == "UNPARSED_NOTICE":
                                log_invoice_ref = proc_invoice_id or "UNMATCHED"
                            else:
                                log_invoice_ref = proc_invoice_id or status
                            log_processed_message(msg_id, log_invoice_ref, received_at=received_at, tenant_id=tenant_id)
                            
                            # Log EMAIL_SENT activity (Bug 12)
                            try:
                                from src.database_sqlite import log_activity
                                _sent_inv = proc_invoice_id or (f"CUSTOMER_REPLIED:{qtn_ref}" if is_customer_reply and qtn_ref else None)
                                log_activity("EMAIL_SENT", invoice_id=_sent_inv,
                                    customer_name=sender_name, customer_email=sender,
                                    description=f"Reply sent to {sender} — Subject: {reply_subject[:80]}",
                                    tenant_id=tenant_id)
                            except Exception:
                                pass
                        
                    print(f"[Success] Processed email from {sender} (status: {status}) and sent reply via SMTP for tenant {tenant_id}.")

                    # --- Handle mixed matched/unmatched items ---
                    if status in ["QUOTE_GENERATED", "QUOTE_UPDATED"]:
                        clean_body_for_deficits = strip_email_history(body)
                        has_attachment = "[From attachment" in clean_body_for_deficits
                        cap_by_stock = False
                        override_qty = None
                        if has_attachment:
                            cap_by_stock = True
                            parts = clean_body_for_deficits.split("[From attachment")
                            email_text_only = parts[0].strip()
                            override_qty = extract_global_quantity_override(email_text_only)
                            body_lines = run_scenario_free(email_text_only, catalog)
                            body_has_products = any(l["matched_sku_id"] != "UNKNOWN" for l in body_lines)
                            if body_has_products or override_qty is not None:
                                cap_by_stock = False

                        matched_lines = run_scenario_free(clean_body_for_deficits, catalog)
                        if override_qty is not None:
                            for line in matched_lines:
                                if line['matched_sku_id'] != "UNKNOWN":
                                    line['quantity'] = override_qty
                        deficit_lines = adjust_quantities_by_stock(matched_lines, catalog, cap_by_stock=cap_by_stock)
                        
                        # Trigger deficit purchase order alert if there are deficits
                        if deficit_lines:
                            print(f"[Deficit PO] {len(deficit_lines)} deficit items found. Sending alert to master.")
                            if master_email and email_user and email_pass:
                                send_deficit_purchase_order_alert(
                                    smtp_server=smtp_server,
                                    smtp_port=smtp_port,
                                    email_user=email_user,
                                    email_pass=email_pass,
                                    master_email=master_email,
                                    customer_name=sender_name,
                                    customer_email=sender,
                                    customer_phone=contact_phone,
                                    original_subject=subject,
                                    deficit_lines=deficit_lines
                                )
                                
                        unmatched_lines = [line for line in matched_lines if line['matched_sku_id'] == "UNKNOWN"]
                        if unmatched_lines:
                            print(f"[Unmatched Products] {len(unmatched_lines)} unmatched items found. Logging & alerting master.")
                            try:
                                from src.database_sqlite import log_unmatched_item
                                unmatched_desc = "\n".join([f"{l['quantity']} x {l['parsed_query']} (original: {l['original_line']})" for l in unmatched_lines])
                                log_unmatched_item(
                                    customer_email=sender,
                                    customer_name=sender_name,
                                    original_body=f"UNMATCHED PRODUCTS REQUESTED:\n{unmatched_desc}\n\nFULL EMAIL BODY:\n{body}",
                                    source="live_email",
                                    tenant_id=tenant_id
                                )
                            except Exception as ue:
                                print(f"[Warning] Failed to log unmatched items: {ue}")

                            if master_email and email_user and email_pass:
                                send_unmatched_products_alert(
                                    smtp_server=smtp_server,
                                    smtp_port=smtp_port,
                                    email_user=email_user,
                                    email_pass=email_pass,
                                    master_email=master_email,
                                    customer_name=sender_name,
                                    customer_email=sender,
                                    customer_phone=contact_phone,
                                    original_subject=subject,
                                    unmatched_lines=unmatched_lines
                                )

                    # --- Handle UNPARSED_NOTICE ---
                    if status == "UNPARSED_NOTICE":
                        print(f"[Unmatched] No valid SKU matches for enquiry from {sender}. Logging & alerting master.")
                        try:
                            from src.database_sqlite import log_unmatched_item
                            log_unmatched_item(
                                customer_email=sender,
                                customer_name=sender_name,
                                original_body=body,
                                source="live_email",
                                tenant_id=tenant_id
                            )
                        except Exception as ue:
                            print(f"[Warning] Failed to log unmatched item: {ue}")

                        if master_email and email_user and email_pass:
                            subject_unmatch_notif = f"[ACTION REQUIRED] Unmatched enquiry from {sender_name} — Manual Quote Needed"
                            body_unmatch_notif = (
                                f"Dear Master User,\n\n"
                                f"The auto-bot received an enquiry but could NOT prepare an automated quote because "
                                f"the requested items do not match any product in our catalogue.\n\n"
                                f"Customer Details:\n"
                                f"- Name: {sender_name}\n"
                                f"- Email: {sender}\n"
                                f"- Contact Number: {contact_phone}\n\n"
                                f"Original Enquiry:\n"
                                f"Subject: {subject}\n"
                                f"--------------------------------------------------\n"
                                f"{body}\n"
                                f"--------------------------------------------------\n\n"
                                f"Please review the above request and prepare a manual quotation directly to the customer.\n\n"
                                f"Regards,\n"
                                f"Trofeo Auto-bot"
                            )
                            send_master_notification(smtp_server, smtp_port, email_user, email_pass, master_email, subject_unmatch_notif, body_unmatch_notif)

                    # 2. Notify Master User of outgoing reply
                    if master_email and email_user and email_pass:
                        subject_reply_notif = f"[Notification] Reply sent to {sender_name}"
                        body_reply_notif = (
                            f"Dear Master User,\n\n"
                            f"The bot has successfully sent a reply to the customer's enquiry.\n\n"
                            f"Customer Details:\n"
                            f"- Name: {sender_name}\n"
                            f"- Email: {sender}\n"
                            f"- Contact Number: {contact_phone}\n\n"
                            f"Reply Details:\n"
                            f"- Subject: {reply_subject}\n"
                            f"- Status: {status}\n\n"
                            f"Replied Message Content:\n"
                            f"--------------------------------------------------\n"
                            f"{plain_body}\n"
                            f"--------------------------------------------------\n\n"
                            f"Regards,\n"
                            f"Trofeo Auto-bot"
                        )
                        send_master_notification(smtp_server, smtp_port, email_user, email_pass, master_email, subject_reply_notif, body_reply_notif)
                    
            # Now, enter native IMAP IDLE to wait for new emails (real-time push)
            print(f"[Email Listener] Entering IDLE state to wait for real-time emails for {tenant_id}...")
            try:
                from src.database_sqlite import update_service_status
                update_service_status("IDLE", tenant_id=tenant_id)
            except Exception:
                pass
            import select
            
            # Send IDLE command
            tag = mail._new_tag().decode()
            mail.send(f'{tag} IDLE\r\n'.encode())
            
            # Read continuation response "+ idling"
            response = mail.readline()
            if b'+' in response:
                sock = mail.socket()
                # Wait up to 2 seconds for socket data (lower timeout to allow iteration over other tenants)
                ready, _, _ = select.select([sock], [], [], 2)
                if ready:
                    print(f"[Email Listener] Real-time email event detected for {tenant_id}!")
                    
            # Send DONE to exit IDLE
            mail.send(b'DONE\r\n')
            # Read tagged response
            mail.readline()
            
            mail.close()
            mail.logout()
            
        except Exception as e:
            print(f"[Email Listener Error] Live processing crashed for tenant {tenant_id}: {e}")
            try:
                from src.database_sqlite import update_service_status
                err_msg = str(e)
                status = "ERROR"
                if "authenticate" in err_msg.lower() or "login" in err_msg.lower() or "credentials" in err_msg.lower() or isinstance(e, imaplib.IMAP4.error):
                    status = "AUTH_FAILED"
                update_service_status(status, error_message=err_msg, tenant_id=tenant_id)
            except Exception:
                pass
